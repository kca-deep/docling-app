"""
채팅 엑셀 내보내기 서비스
LLM 응답 데이터를 엑셀 파일로 변환합니다.
마크다운 형식의 콘텐츠를 구조화된 엑셀로 변환합니다.
"""

import logging
import re
import unicodedata
from io import BytesIO
from typing import List, Tuple, Optional, Dict, Any
from datetime import datetime
from dataclasses import dataclass
from enum import Enum

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from backend.services.tool_executor_service import (
    ToolResult,
    file_storage,
    FileSizeExceededError,
    StorageCapacityExceededError
)
from backend.utils.text_normalizer import normalize_spaces

logger = logging.getLogger("uvicorn")


class RowType(Enum):
    """엑셀 행 타입"""
    HEADING1 = "h1"
    HEADING2 = "h2"
    HEADING3 = "h3"
    HEADING4 = "h4"
    LIST_ITEM = "list"
    NUMBERED_LIST = "numbered"
    TABLE_HEADER = "table_header"
    TABLE_ROW = "table_row"
    PARAGRAPH = "paragraph"
    EMPTY = "empty"
    CODE_BLOCK = "code"
    BLOCKQUOTE = "quote"


@dataclass
class ParsedRow:
    """파싱된 행 데이터"""
    row_type: RowType
    content: str
    level: int = 0  # 들여쓰기 레벨 (목록용)
    cells: List[str] = None  # 테이블 셀 (테이블용)


class ChatExcelExportService:
    """
    채팅 데이터를 엑셀로 내보내는 서비스
    마크다운 형식의 콘텐츠를 구조화된 엑셀로 변환합니다.
    """

    # 스타일 상수 - 제목별 차별화
    H1_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    H1_FONT = Font(bold=True, color="FFFFFF", size=14)

    H2_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    H2_FONT = Font(bold=True, color="FFFFFF", size=12)

    H3_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    H3_FONT = Font(bold=True, color="000000", size=11)

    H4_FILL = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    H4_FONT = Font(bold=True, color="000000", size=10)

    # 기존 스타일 (테이블 헤더용)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    # 코드 블록 스타일
    CODE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    CODE_FONT = Font(name="Consolas", size=10)

    # 인용 스타일
    QUOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    QUOTE_FONT = Font(italic=True, color="666666")

    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    def __init__(self):
        pass

    def _create_rich_text(self, text: str) -> CellRichText:
        """
        마크다운 인라인 서식을 CellRichText로 변환

        지원 서식:
        - **bold** → 볼드
        - *italic* → 이탤릭
        - `code` → 코드 (Consolas 폰트)

        Args:
            text: 마크다운 인라인 서식이 포함된 텍스트

        Returns:
            CellRichText: 서식이 적용된 리치 텍스트
        """
        # 패턴: **bold**, *italic*, `code`
        pattern = r'(\*\*(.+?)\*\*|\*(.+?)\*|`([^`]+)`)'

        parts = []
        last_end = 0

        for match in re.finditer(pattern, text):
            # 매치 전 일반 텍스트
            if match.start() > last_end:
                parts.append(text[last_end:match.start()])

            full_match = match.group(0)

            if full_match.startswith('**'):
                # Bold
                bold_font = InlineFont(b=True)
                parts.append(TextBlock(bold_font, match.group(2)))
            elif full_match.startswith('`'):
                # Code - 백그라운드 색상은 CellRichText에서 지원 안됨, 폰트만 적용
                code_font = InlineFont(rFont="Consolas")
                parts.append(TextBlock(code_font, match.group(4)))
            elif full_match.startswith('*'):
                # Italic
                italic_font = InlineFont(i=True)
                parts.append(TextBlock(italic_font, match.group(3)))

            last_end = match.end()

        # 나머지 텍스트
        if last_end < len(text):
            parts.append(text[last_end:])

        # 서식이 없으면 일반 문자열 반환
        if not parts:
            return text

        # 모든 요소가 문자열이면 일반 문자열 반환
        if all(isinstance(p, str) for p in parts):
            return text

        return CellRichText(*parts)

    def _has_inline_formatting(self, text: str) -> bool:
        """텍스트에 인라인 마크다운 서식이 있는지 확인"""
        return bool(re.search(r'\*\*(.+?)\*\*|\*(.+?)\*|`([^`]+)`', text))

    def _parse_csv_data(self, data: str) -> List[List[str]]:
        """CSV 형식 데이터 파싱"""
        rows = []
        for line in data.strip().split('\n'):
            if line.strip():
                # 쉼표로 분리하되, 따옴표 내 쉼표는 보존
                cells = []
                current = ""
                in_quotes = False
                for char in line:
                    if char == '"':
                        in_quotes = not in_quotes
                    elif char == ',' and not in_quotes:
                        cells.append(current.strip().strip('"'))
                        current = ""
                    else:
                        current += char
                cells.append(current.strip().strip('"'))
                rows.append(cells)
        return rows

    def _parse_markdown_table(self, data: str) -> List[List[str]]:
        """마크다운 테이블 형식 파싱 (| 구분)"""
        rows = []
        for line in data.strip().split('\n'):
            line = line.strip()
            if not line or line.startswith('|--') or line.startswith('| --') or re.match(r'^\|[\s\-:]+\|$', line):
                # 구분선 스킵
                continue
            if '|' in line:
                # | 로 분리
                cells = [cell.strip() for cell in line.split('|')]
                # 앞뒤 빈 셀 제거
                cells = [c for c in cells if c or cells.index(c) not in [0, len(cells)-1]]
                if cells:
                    rows.append(cells)
        return rows

    def _parse_line_data(self, data: str) -> List[List[str]]:
        """줄바꿈 구분 데이터 파싱 (단일 컬럼)"""
        rows = []
        for line in data.strip().split('\n'):
            if line.strip():
                rows.append([line.strip()])
        return rows

    def _parse_markdown_content(self, data: str) -> List[ParsedRow]:
        """
        마크다운 콘텐츠를 구조화된 행 데이터로 파싱

        Args:
            data: 마크다운 형식의 텍스트

        Returns:
            List[ParsedRow]: 파싱된 행 리스트
        """
        rows = []
        lines = data.strip().split('\n')
        in_code_block = False
        code_block_content = []
        in_table = False
        table_header_processed = False

        for line in lines:
            stripped = line.strip()

            # 코드 블록 처리
            if stripped.startswith('```'):
                if in_code_block:
                    # 코드 블록 종료
                    if code_block_content:
                        rows.append(ParsedRow(
                            row_type=RowType.CODE_BLOCK,
                            content='\n'.join(code_block_content)
                        ))
                    code_block_content = []
                    in_code_block = False
                else:
                    # 코드 블록 시작
                    in_code_block = True
                continue

            if in_code_block:
                code_block_content.append(line)
                continue

            # 빈 줄 처리 (연속 빈 줄은 하나로 병합)
            if not stripped:
                in_table = False
                table_header_processed = False
                # 이전 행이 빈 행이 아닐 때만 추가
                if not rows or rows[-1].row_type != RowType.EMPTY:
                    rows.append(ParsedRow(row_type=RowType.EMPTY, content=""))
                continue

            # 테이블 행 감지
            if '|' in stripped and stripped.count('|') >= 2:
                cells = [c.strip() for c in stripped.split('|')]
                cells = [c for i, c in enumerate(cells) if c or (i != 0 and i != len(cells) - 1)]

                # 테이블 구분선 스킵 (모든 셀이 -와 :로만 구성)
                if cells and all(re.match(r'^[\-:]+$', cell) for cell in cells):
                    continue

                if not in_table:
                    # 테이블 헤더
                    in_table = True
                    table_header_processed = True
                    rows.append(ParsedRow(
                        row_type=RowType.TABLE_HEADER,
                        content=stripped,
                        cells=cells
                    ))
                else:
                    # 테이블 데이터 행
                    rows.append(ParsedRow(
                        row_type=RowType.TABLE_ROW,
                        content=stripped,
                        cells=cells
                    ))
                continue

            in_table = False
            table_header_processed = False

            # 제목 감지 (# ~ ####)
            heading_match = re.match(r'^(#{1,4})\s+(.+)$', stripped)
            if heading_match:
                level = len(heading_match.group(1))
                content = heading_match.group(2).strip()
                row_type = {
                    1: RowType.HEADING1,
                    2: RowType.HEADING2,
                    3: RowType.HEADING3,
                    4: RowType.HEADING4
                }.get(level, RowType.HEADING4)
                rows.append(ParsedRow(row_type=row_type, content=content))
                continue

            # 인용 감지
            if stripped.startswith('>'):
                content = stripped.lstrip('>').strip()
                rows.append(ParsedRow(row_type=RowType.BLOCKQUOTE, content=content))
                continue

            # 순서 없는 목록 감지 (-, *, +)
            list_match = re.match(r'^(\s*)([-*+])\s+(.+)$', line)
            if list_match:
                indent = len(list_match.group(1))
                level = indent // 2  # 2칸 들여쓰기당 1레벨
                content = list_match.group(3).strip()
                rows.append(ParsedRow(
                    row_type=RowType.LIST_ITEM,
                    content=content,
                    level=level
                ))
                continue

            # 순서 있는 목록 감지 (1. 2. 등)
            numbered_match = re.match(r'^(\s*)(\d+)\.\s+(.+)$', line)
            if numbered_match:
                indent = len(numbered_match.group(1))
                level = indent // 2
                number = numbered_match.group(2)
                content = numbered_match.group(3).strip()
                rows.append(ParsedRow(
                    row_type=RowType.NUMBERED_LIST,
                    content=f"{number}. {content}",
                    level=level
                ))
                continue

            # 일반 단락 - 인라인 서식 보존 (CellRichText로 변환)
            rows.append(ParsedRow(row_type=RowType.PARAGRAPH, content=stripped))

        # 코드 블록이 닫히지 않은 경우 처리
        if code_block_content:
            rows.append(ParsedRow(
                row_type=RowType.CODE_BLOCK,
                content='\n'.join(code_block_content)
            ))

        return rows

    def _clean_markdown_formatting(self, text: str) -> str:
        """마크다운 인라인 서식을 정리 (볼드, 이탤릭 등)"""
        # **bold** → bold
        text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
        # *italic* → italic
        text = re.sub(r'\*(.+?)\*', r'\1', text)
        # __bold__ → bold
        text = re.sub(r'__(.+?)__', r'\1', text)
        # _italic_ → italic
        text = re.sub(r'_(.+?)_', r'\1', text)
        # `code` → code
        text = re.sub(r'`(.+?)`', r'\1', text)
        # [link](url) → link
        text = re.sub(r'\[(.+?)\]\(.+?\)', r'\1', text)
        return text

    def _is_markdown_content(self, data: str) -> bool:
        """
        데이터가 마크다운 형식인지 감지

        Returns:
            bool: 마크다운 형식이면 True
        """
        # 마크다운 패턴 감지
        markdown_patterns = [
            r'^#{1,4}\s+',      # 제목
            r'^[-*+]\s+',       # 순서 없는 목록
            r'^\d+\.\s+',       # 순서 있는 목록
            r'^>\s*',           # 인용
            r'```',             # 코드 블록
            r'\*\*.+?\*\*',     # 볼드
            r'\[.+?\]\(.+?\)',  # 링크
        ]

        lines = data.strip().split('\n')
        markdown_line_count = 0

        for line in lines:
            for pattern in markdown_patterns:
                if re.search(pattern, line):
                    markdown_line_count += 1
                    break

        # 전체 라인의 20% 이상이 마크다운 패턴을 포함하면 마크다운으로 간주
        return markdown_line_count >= len(lines) * 0.2 if lines else False

    def _detect_and_parse(self, data: str) -> Tuple[List[List[str]], str]:
        """
        데이터 형식 자동 감지 및 파싱

        Returns:
            Tuple[rows, format_type]
        """
        data = data.strip()

        # 마크다운 콘텐츠 감지 (우선순위 높음)
        if self._is_markdown_content(data):
            return None, "markdown"  # 마크다운은 별도 처리

        # 마크다운 테이블 감지
        if '|' in data and data.count('|') >= 2:
            rows = self._parse_markdown_table(data)
            if rows and len(rows[0]) > 1:
                return rows, "markdown_table"

        # CSV 형식 감지
        lines = data.split('\n')
        if len(lines) > 0 and ',' in lines[0]:
            rows = self._parse_csv_data(data)
            if rows and len(rows[0]) > 1:
                return rows, "csv"

        # 기본: 줄바꿈 구분
        rows = self._parse_line_data(data)
        return rows, "lines"

    def _apply_header_style(self, ws, row_num: int, col_count: int):
        """헤더 행 스타일 적용"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

    def _apply_cell_style(self, ws, row_num: int, col_count: int):
        """일반 셀 스타일 적용"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = self.BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    def _get_display_width(self, text: str) -> int:
        """
        텍스트의 디스플레이 너비 계산

        한글 및 동아시아 문자(CJK)는 2칸, ASCII는 1칸으로 계산
        """
        if not text:
            return 0
        width = 0
        for char in str(text):
            ea_width = unicodedata.east_asian_width(char)
            if ea_width in ('W', 'F'):
                width += 2
            else:
                width += 1
        return width

    def _auto_adjust_column_width(self, ws, table_columns: List[int] = None):
        """
        컬럼 너비 자동 조정

        Args:
            ws: 워크시트
            table_columns: 테이블 컬럼 인덱스 리스트 (균등 배분용)
        """
        column_widths = {}

        for column in ws.columns:
            max_width = 0
            column_idx = column[0].column
            column_letter = get_column_letter(column_idx)

            for cell in column:
                try:
                    if cell.value:
                        cell_width = self._get_display_width(cell.value)
                        if cell_width > max_width:
                            max_width = cell_width
                except:
                    pass

            # 여유 공간 추가 및 최대/최소 제한
            adjusted_width = min(max_width + 2, 60)  # 최대 60
            adjusted_width = max(adjusted_width, 8)   # 최소 8
            column_widths[column_letter] = adjusted_width

        # 테이블 컬럼 균등 배분 (선택사항)
        if table_columns and len(table_columns) > 1:
            total_width = sum(column_widths.get(get_column_letter(c), 10) for c in table_columns)
            avg_width = total_width / len(table_columns)

            # 너무 좁은 컬럼은 평균에 가깝게 조정
            for col_idx in table_columns:
                col_letter = get_column_letter(col_idx)
                if col_letter in column_widths:
                    current = column_widths[col_letter]
                    # 평균의 50% 미만이면 평균의 70%로 조정
                    if current < avg_width * 0.5:
                        column_widths[col_letter] = max(current, avg_width * 0.7)

        # 컬럼 너비 적용
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    def _apply_row_style(self, ws, row_num: int, parsed_row: ParsedRow, col_count: int = 1):
        """
        ParsedRow 타입에 따라 적절한 스타일 적용

        Args:
            ws: 워크시트
            row_num: 행 번호
            parsed_row: 파싱된 행 데이터
            col_count: 컬럼 수 (테이블용)
        """
        cell = ws.cell(row=row_num, column=1)

        if parsed_row.row_type == RowType.HEADING1:
            cell.fill = self.H1_FILL
            cell.font = self.H1_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif parsed_row.row_type == RowType.HEADING2:
            cell.fill = self.H2_FILL
            cell.font = self.H2_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif parsed_row.row_type == RowType.HEADING3:
            cell.fill = self.H3_FILL
            cell.font = self.H3_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif parsed_row.row_type == RowType.HEADING4:
            cell.fill = self.H4_FILL
            cell.font = self.H4_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif parsed_row.row_type == RowType.CODE_BLOCK:
            cell.fill = self.CODE_FILL
            cell.font = self.CODE_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        elif parsed_row.row_type == RowType.BLOCKQUOTE:
            cell.fill = self.QUOTE_FILL
            cell.font = self.QUOTE_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True, indent=1)
        elif parsed_row.row_type in (RowType.LIST_ITEM, RowType.NUMBERED_LIST):
            cell.alignment = Alignment(vertical='center', wrap_text=True, indent=parsed_row.level + 1)
        elif parsed_row.row_type == RowType.TABLE_HEADER:
            # 테이블 헤더는 각 셀에 스타일 적용
            for col_idx in range(1, col_count + 1):
                tc = ws.cell(row=row_num, column=col_idx)
                tc.fill = self.HEADER_FILL
                tc.font = self.HEADER_FONT
                tc.border = self.BORDER
                tc.alignment = Alignment(horizontal='center', vertical='center')
        elif parsed_row.row_type == RowType.TABLE_ROW:
            # 테이블 데이터 행
            for col_idx in range(1, col_count + 1):
                tc = ws.cell(row=row_num, column=col_idx)
                tc.border = self.THIN_BORDER
                tc.alignment = Alignment(vertical='center', wrap_text=True)
        else:
            # 일반 단락
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    def _export_markdown_to_excel(self, ws, data: str):
        """
        마크다운 콘텐츠를 구조화된 엑셀로 변환

        Args:
            ws: 워크시트
            data: 마크다운 콘텐츠
        """
        parsed_rows = self._parse_markdown_content(data)
        logger.info(f"[ExcelExport] Parsed markdown: {len(parsed_rows)} rows")

        if not parsed_rows:
            ws.cell(row=1, column=1, value="데이터가 없습니다.")
            return

        excel_row = 1
        max_cols = 1
        table_columns = set()  # 테이블에 사용된 컬럼 추적

        for parsed_row in parsed_rows:
            # 빈 줄은 빈 행으로 처리 (연속된 빈 줄은 하나로)
            if parsed_row.row_type == RowType.EMPTY:
                excel_row += 1
                continue

            # 테이블 행 처리
            if parsed_row.row_type in (RowType.TABLE_HEADER, RowType.TABLE_ROW):
                if parsed_row.cells:
                    col_count = len(parsed_row.cells)
                    max_cols = max(max_cols, col_count)
                    # 테이블 컬럼 인덱스 추적
                    for col_idx in range(1, col_count + 1):
                        table_columns.add(col_idx)

                    for col_idx, cell_value in enumerate(parsed_row.cells, 1):
                        ws.cell(row=excel_row, column=col_idx, value=cell_value)

                    self._apply_row_style(ws, excel_row, parsed_row, col_count)
            else:
                # 목록 항목에 들여쓰기 표시 추가
                if parsed_row.row_type == RowType.LIST_ITEM:
                    indent = "  " * parsed_row.level
                    prefix = f"{indent}• "
                    raw_content = parsed_row.content
                elif parsed_row.row_type == RowType.NUMBERED_LIST:
                    indent = "  " * parsed_row.level
                    prefix = f"{indent}"
                    raw_content = parsed_row.content
                else:
                    prefix = ""
                    raw_content = parsed_row.content

                # 인라인 서식 적용
                if self._has_inline_formatting(raw_content):
                    # CellRichText로 서식 적용
                    rich_content = self._create_rich_text(raw_content)
                    if isinstance(rich_content, CellRichText) and prefix:
                        # prefix와 rich_content 결합
                        combined_parts = [prefix]
                        # CellRichText에서 직접 요소 추출 (as_list()는 문자열만 반환)
                        for elem in rich_content:
                            combined_parts.append(elem)
                        cell_value = CellRichText(*combined_parts)
                    elif isinstance(rich_content, CellRichText):
                        cell_value = rich_content
                    else:
                        cell_value = prefix + raw_content
                else:
                    cell_value = prefix + raw_content

                ws.cell(row=excel_row, column=1, value=cell_value)
                self._apply_row_style(ws, excel_row, parsed_row)

            excel_row += 1

        # 컬럼 너비 자동 조정 (테이블 컬럼 정보 전달)
        self._auto_adjust_column_width(ws, list(table_columns) if table_columns else None)

        # 첫 번째 열 너비를 충분히 확보 (마크다운 콘텐츠용)
        if max_cols == 1:
            ws.column_dimensions['A'].width = 80

    def export_to_excel(
        self,
        data: str,
        filename: str = "export",
        sheet_name: str = "Sheet1"
    ) -> bytes:
        """
        데이터를 엑셀 파일로 변환

        Args:
            data: 변환할 데이터 (마크다운, CSV, 테이블, 또는 일반 텍스트)
            filename: 파일명 (확장자 제외)
            sheet_name: 시트 이름

        Returns:
            bytes: 엑셀 파일 바이너리
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 특수 공백 정규화
        data = normalize_spaces(data)

        # 데이터 파싱
        rows, format_type = self._detect_and_parse(data)
        logger.info(f"[ExcelExport] Detected format: {format_type}")

        # 마크다운 형식은 별도 처리
        if format_type == "markdown":
            self._export_markdown_to_excel(ws, data)
        elif not rows:
            # 빈 데이터인 경우 기본 메시지
            ws.cell(row=1, column=1, value="데이터가 없습니다.")
        else:
            logger.info(f"[ExcelExport] Processing {len(rows)} rows as {format_type}")
            col_count = max(len(row) for row in rows)

            for row_idx, row_data in enumerate(rows, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

                # 첫 번째 행은 헤더로 스타일링
                if row_idx == 1:
                    self._apply_header_style(ws, row_idx, col_count)
                else:
                    self._apply_cell_style(ws, row_idx, col_count)

            # 컬럼 너비 자동 조정
            self._auto_adjust_column_width(ws)

        # 바이너리로 저장
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    async def handle_export_to_excel(
        self,
        tool_call_id: str,
        arguments: dict
    ) -> ToolResult:
        """
        export_to_excel 도구 핸들러

        Args:
            tool_call_id: 도구 호출 ID
            arguments: 도구 인자
                - data: 내보낼 데이터
                - filename: 파일명 (선택)
                - sheet_name: 시트 이름 (선택)

        Returns:
            ToolResult: 실행 결과
        """
        try:
            data = arguments.get("data", "")
            if not data:
                return ToolResult(
                    tool_call_id=tool_call_id,
                    tool_name="export_to_excel",
                    success=False,
                    action_type="message",
                    error="내보낼 데이터가 없습니다."
                )

            filename = arguments.get("filename", "export")
            sheet_name = arguments.get("sheet_name", "Sheet1")

            # 엑셀 생성
            excel_bytes = self.export_to_excel(data, filename, sheet_name)

            # 파일 저장소에 저장 (중복 확장자 방지)
            if filename.lower().endswith('.xlsx'):
                full_filename = filename
            else:
                full_filename = f"{filename}.xlsx"
            file_id = file_storage.store(
                filename=full_filename,
                content=excel_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            return ToolResult(
                tool_call_id=tool_call_id,
                tool_name="export_to_excel",
                success=True,
                action_type="download",
                file_id=file_id,
                filename=full_filename,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                message=f"엑셀 파일 '{full_filename}'이(가) 생성되었습니다."
            )

        except FileSizeExceededError as e:
            logger.warning(f"[ExcelExport] File size exceeded: {e}")
            return ToolResult(
                tool_call_id=tool_call_id,
                tool_name="export_to_excel",
                success=False,
                action_type="message",
                error=str(e)
            )

        except StorageCapacityExceededError as e:
            logger.warning(f"[ExcelExport] Storage capacity exceeded: {e}")
            return ToolResult(
                tool_call_id=tool_call_id,
                tool_name="export_to_excel",
                success=False,
                action_type="message",
                error=str(e)
            )

        except Exception as e:
            logger.error(f"[ExcelExport] Export failed: {e}")
            return ToolResult(
                tool_call_id=tool_call_id,
                tool_name="export_to_excel",
                success=False,
                action_type="message",
                error=f"엑셀 파일 생성 중 오류가 발생했습니다: {str(e)}"
            )


# 싱글톤 인스턴스
chat_excel_export_service = ChatExcelExportService()
