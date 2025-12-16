"""
Excel 내보내기 서비스
셀프진단 결과를 Excel 파일로 내보내기
"""
import logging
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.models.schemas import SelfCheckDetailResponse

logger = logging.getLogger(__name__)


class ExcelExportService:
    """셀프진단 결과 Excel 내보내기 서비스"""

    def __init__(self):
        # 스타일 정의
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self.cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.center_alignment = Alignment(horizontal="center", vertical="center")

        self.thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # 검토 대상 여부에 따른 배경색
        self.review_yes_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
        self.review_no_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")

    def _answer_to_korean(self, answer: str | None) -> str:
        """답변 값을 한국어로 변환"""
        mapping = {
            "yes": "예",
            "no": "아니오",
            "unknown": "모름",
            "need_check": "확인필요"
        }
        return mapping.get(answer, "-") if answer else "-"

    def _match_status_to_korean(self, status: str) -> str:
        """일치 상태 한국어 변환"""
        mapping = {
            "match": "일치",
            "mismatch": "불일치",
            "reference": "AI참조",
            "keep": "유지"
        }
        return mapping.get(status, status)

    def _set_column_widths(self, ws, widths: List[int]):
        """컬럼 너비 설정"""
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _add_header_row(self, ws, headers: List[str], row: int = 1):
        """헤더 행 추가"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

    async def export_selfcheck_excel(
        self,
        submissions: List[SelfCheckDetailResponse]
    ) -> bytes:
        """
        셀프진단 결과를 Excel로 내보내기

        Args:
            submissions: 셀프진단 상세 정보 목록

        Returns:
            bytes: Excel 파일 바이트
        """
        wb = Workbook()

        # === 시트 1: 진단요약 ===
        ws_summary = wb.active
        ws_summary.title = "진단요약"

        summary_headers = [
            "No", "과제명", "담당부서", "담당자", "연락처", "이메일",
            "검토대상", "검토사유", "진단일시", "사용모델"
        ]
        self._add_header_row(ws_summary, summary_headers)
        self._set_column_widths(ws_summary, [5, 40, 15, 10, 15, 25, 10, 30, 20, 15])

        for idx, sub in enumerate(submissions, 1):
            row = idx + 1
            ws_summary.cell(row=row, column=1, value=idx).alignment = self.center_alignment
            ws_summary.cell(row=row, column=2, value=sub.project_name).alignment = self.cell_alignment
            ws_summary.cell(row=row, column=3, value=sub.department).alignment = self.cell_alignment
            ws_summary.cell(row=row, column=4, value=sub.manager_name).alignment = self.cell_alignment
            ws_summary.cell(row=row, column=5, value=sub.contact or "-").alignment = self.cell_alignment
            ws_summary.cell(row=row, column=6, value=sub.email or "-").alignment = self.cell_alignment

            review_cell = ws_summary.cell(row=row, column=7, value="예" if sub.requires_review else "아니오")
            review_cell.alignment = self.center_alignment
            review_cell.fill = self.review_yes_fill if sub.requires_review else self.review_no_fill

            ws_summary.cell(row=row, column=8, value=sub.review_reason or "-").alignment = self.cell_alignment
            ws_summary.cell(row=row, column=9, value=sub.created_at[:19].replace("T", " ") if sub.created_at else "-").alignment = self.center_alignment
            ws_summary.cell(row=row, column=10, value=sub.used_model or "-").alignment = self.cell_alignment

            # 테두리 적용
            for col in range(1, 11):
                ws_summary.cell(row=row, column=col).border = self.thin_border

        # === 시트 2: 점검항목상세 ===
        ws_items = wb.create_sheet("점검항목상세")

        items_headers = [
            "과제명", "항목번호", "카테고리", "항목명", "사용자선택",
            "AI분석", "일치여부", "신뢰도", "AI근거"
        ]
        self._add_header_row(ws_items, items_headers)
        self._set_column_widths(ws_items, [30, 10, 10, 20, 10, 10, 10, 10, 50])

        items_row = 2
        for sub in submissions:
            for item in sub.items:
                ws_items.cell(row=items_row, column=1, value=sub.project_name).alignment = self.cell_alignment
                ws_items.cell(row=items_row, column=2, value=item.item_number).alignment = self.center_alignment
                ws_items.cell(row=items_row, column=3, value="필수" if item.item_category == "required" else "선택").alignment = self.center_alignment
                ws_items.cell(row=items_row, column=4, value=item.short_label).alignment = self.cell_alignment
                ws_items.cell(row=items_row, column=5, value=self._answer_to_korean(item.user_answer)).alignment = self.center_alignment
                ws_items.cell(row=items_row, column=6, value=self._answer_to_korean(item.llm_answer)).alignment = self.center_alignment
                ws_items.cell(row=items_row, column=7, value=self._match_status_to_korean(item.match_status)).alignment = self.center_alignment
                ws_items.cell(row=items_row, column=8, value=f"{int(item.llm_confidence * 100)}%").alignment = self.center_alignment
                ws_items.cell(row=items_row, column=9, value=item.llm_evidence or "-").alignment = self.cell_alignment

                # 테두리 적용
                for col in range(1, 10):
                    ws_items.cell(row=items_row, column=col).border = self.thin_border

                items_row += 1

        # === 시트 3: 유사과제 ===
        ws_similar = wb.create_sheet("유사과제")

        similar_headers = [
            "원과제명", "유사과제명", "부서", "담당자", "유사도(%)", "유사사유", "등록일"
        ]
        self._add_header_row(ws_similar, similar_headers)
        self._set_column_widths(ws_similar, [30, 30, 15, 10, 10, 40, 15])

        similar_row = 2
        for sub in submissions:
            for sp in sub.similar_projects:
                ws_similar.cell(row=similar_row, column=1, value=sub.project_name).alignment = self.cell_alignment
                ws_similar.cell(row=similar_row, column=2, value=sp.project_name).alignment = self.cell_alignment
                ws_similar.cell(row=similar_row, column=3, value=sp.department).alignment = self.cell_alignment
                ws_similar.cell(row=similar_row, column=4, value=sp.manager_name).alignment = self.cell_alignment
                ws_similar.cell(row=similar_row, column=5, value=sp.similarity_score).alignment = self.center_alignment
                ws_similar.cell(row=similar_row, column=6, value=sp.similarity_reason).alignment = self.cell_alignment
                ws_similar.cell(row=similar_row, column=7, value=sp.created_at[:10] if sp.created_at else "-").alignment = self.center_alignment

                # 테두리 적용
                for col in range(1, 8):
                    ws_similar.cell(row=similar_row, column=col).border = self.thin_border

                similar_row += 1

        # 유사과제가 없는 경우 안내 메시지
        if similar_row == 2:
            ws_similar.cell(row=2, column=1, value="유사 과제가 없습니다.").alignment = self.cell_alignment

        # Excel 파일로 저장
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


# 싱글톤 인스턴스
excel_export_service = ExcelExportService()
