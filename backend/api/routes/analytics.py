"""
Analytics API 라우터
채팅 로그 및 통계 조회 엔드포인트
인증 필수: 관리자만 접근 가능
"""

import logging
import pandas as pd
import numpy as np
import io
from urllib.parse import quote
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, HTTPException, Depends, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, desc
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell import WriteOnlyCell

from backend.database import get_db
from backend.dependencies.auth import get_current_active_user
from backend.models.chat_session import ChatSession
from backend.utils.timezone import now_naive
from backend.models.chat_statistics import ChatStatistics
from backend.services.statistics_service import statistics_service
from backend.services.conversation_service import conversation_service
from backend.services.hybrid_logging_service import hybrid_logging_service

# 로거 설정
logger = logging.getLogger(__name__)


# 공통 정규화 유틸리티 사용
from backend.utils.normalize import normalize_collection as _normalize_collection

router = APIRouter(
    prefix="/api/analytics",
    tags=["analytics"],
    dependencies=[Depends(get_current_active_user)]  # 모든 엔드포인트 인증 필수
)


@router.get("/summary")
async def get_analytics_summary(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름"),
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    db: Session = Depends(get_db)
):
    """
    컬렉션별 통계 요약 조회

    Args:
        collection_name: 특정 컬렉션 필터 (선택)
        date_from: 시작 날짜 (선택)
        date_to: 종료 날짜 (선택)
        db: 데이터베이스 세션

    Returns:
        dict: 통계 요약
            - total_queries: 총 쿼리 수
            - unique_sessions: 고유 세션 수
            - total_tokens: 총 토큰 사용량
            - error_count: 에러 횟수
            - avg_response_time_ms: 평균 응답 시간
            - period: 조회 기간 정보
            - collections: 컬렉션 목록
            - top_queries: 인기 쿼리 목록
    """
    try:
        # "ALL" → None 정규화 (전체 조회)
        normalized_collection = _normalize_collection(collection_name)

        summary = await statistics_service.get_summary(
            collection_name=normalized_collection,
            date_from=date_from,
            date_to=date_to,
            db=db
        )
        return summary

    except Exception as e:
        logger.error(f"통계 요약 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=f"통계 요약 조회 실패: {str(e)}")


@router.get("/timeline")
async def get_timeline_data(
    collection_name: str = Query(..., description="컬렉션 이름 (전체: ALL)"),
    period: str = Query("daily", description="집계 주기 (daily/hourly)"),
    days: int = Query(7, description="최근 N일간 데이터"),
    db: Session = Depends(get_db)
):
    """
    시계열 데이터 조회

    Args:
        collection_name: 컬렉션 이름 ('ALL'은 전체 조회)
        period: 집계 주기 ("daily" 또는 "hourly")
        days: 조회할 일수 (기본: 7일)
        db: 데이터베이스 세션

    Returns:
        list: 시계열 데이터
            - date: 날짜
            - hour: 시간 (hourly인 경우)
            - queries: 쿼리 수
            - sessions: 세션 수
            - avg_response_time: 평균 응답 시간
            - errors: 에러 수
    """
    try:
        if period not in ["daily", "hourly"]:
            raise ValueError(f"Invalid period: {period}")

        if days < 1 or days > 90:
            raise ValueError(f"Days must be between 1 and 90")

        # "ALL" → None 정규화 (전체 조회)
        normalized_collection = _normalize_collection(collection_name)

        timeline = await statistics_service.get_timeline(
            collection_name=normalized_collection,
            period=period,
            days=days,
            db=db
        )
        return {"data": timeline, "collection_name": collection_name, "period": period}

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"타임라인 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=f"타임라인 조회 실패: {str(e)}")


@router.get("/sessions")
async def get_sessions(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름"),
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    has_error: Optional[bool] = Query(None, description="에러 포함 여부"),
    skip: int = Query(0, ge=0, description="건너뛸 항목 수"),
    limit: int = Query(20, ge=1, le=100, description="조회할 항목 수"),
    db: Session = Depends(get_db)
):
    """
    채팅 세션 목록 조회

    Args:
        collection_name: 컬렉션 필터
        date_from: 시작 날짜
        date_to: 종료 날짜
        has_error: 에러 세션만 필터
        skip: 페이지네이션 오프셋
        limit: 조회 개수
        db: 데이터베이스 세션

    Returns:
        dict: 세션 목록과 메타데이터
            - sessions: 세션 리스트
            - total: 전체 개수
            - skip: 오프셋
            - limit: 조회 개수
    """
    try:
        query = db.query(ChatSession)

        # 필터 적용
        if collection_name:
            query = query.filter(ChatSession.collection_name == collection_name)
        if date_from:
            query = query.filter(ChatSession.started_at >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            query = query.filter(ChatSession.started_at <= datetime.combine(date_to, datetime.max.time()))
        if has_error is not None:
            query = query.filter(ChatSession.has_error == (1 if has_error else 0))

        # 전체 카운트
        total = query.count()

        # 페이지네이션과 정렬
        sessions = query.order_by(desc(ChatSession.started_at)).offset(skip).limit(limit).all()

        return {
            "sessions": [session.to_dict() for session in sessions],
            "total": total,
            "skip": skip,
            "limit": limit
        }

    except Exception as e:
        logger.error(f"세션 목록 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=f"세션 목록 조회 실패: {str(e)}")


@router.get("/conversations")
async def get_conversations(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름"),
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    limit: int = Query(20, ge=1, le=100, description="조회할 항목 수")
):
    """
    저장된 대화 내역 조회 (관리자용)

    Args:
        collection_name: 컬렉션 필터
        date_from: 시작 날짜
        date_to: 종료 날짜
        limit: 조회 개수

    Returns:
        dict: 대화 목록
            - conversations: 대화 리스트
            - total: 조회된 개수
    """
    try:
        # 날짜 변환
        start_date = datetime.combine(date_from, datetime.min.time()) if date_from else None
        end_date = datetime.combine(date_to, datetime.max.time()) if date_to else None

        conversations = await conversation_service.read_conversations(
            start_date=start_date,
            end_date=end_date,
            collection_name=collection_name,
            limit=limit
        )

        return {
            "conversations": conversations,
            "total": len(conversations)
        }

    except Exception as e:
        logger.error(f"대화 내역 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=f"대화 내역 조회 실패: {str(e)}")


@router.get("/logs")
async def get_logs(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름"),
    session_id: Optional[str] = Query(None, description="세션 ID"),
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    limit: int = Query(100, ge=1, le=1000, description="조회할 항목 수")
):
    """
    원시 로그 데이터 조회

    Args:
        collection_name: 컬렉션 필터
        session_id: 세션 ID 필터
        date_from: 시작 날짜
        date_to: 종료 날짜
        limit: 조회 개수

    Returns:
        dict: 로그 데이터
            - logs: 로그 리스트
            - total: 조회된 개수
    """
    try:
        # 날짜 변환
        start_date = datetime.combine(date_from, datetime.min.time()) if date_from else None
        end_date = datetime.combine(date_to, datetime.max.time()) if date_to else None

        logs = await hybrid_logging_service.read_logs(
            start_date=start_date,
            end_date=end_date,
            collection_name=collection_name,
            session_id=session_id,
            limit=limit
        )

        return {
            "logs": logs,
            "total": len(logs)
        }

    except Exception as e:
        logger.error(f"로그 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=f"로그 조회 실패: {str(e)}")


@router.post("/aggregate")
async def trigger_aggregation(
    background_tasks: BackgroundTasks,
    target_date: date = Query(..., description="집계할 날짜")
):
    """
    특정 날짜의 통계 집계 트리거 (관리자용)

    Args:
        target_date: 집계할 날짜
        background_tasks: 백그라운드 태스크

    Returns:
        dict: 집계 작업 상태
    """
    try:
        # 백그라운드로 집계 실행
        async def run_aggregation():
            # 백그라운드 태스크에서 새 DB 세션 생성 (요청 컨텍스트 외부에서 실행되므로)
            from backend.database import SessionLocal
            db = SessionLocal()
            try:
                result = await statistics_service.aggregate_daily_stats(target_date, db)
                logger.info(f"집계 완료: {result}")
            finally:
                db.close()

        background_tasks.add_task(run_aggregation)

        return {
            "message": f"통계 집계 작업이 시작되었습니다: {target_date}",
            "date": target_date.isoformat(),
            "status": "started"
        }

    except Exception as e:
        logger.error(f"집계 트리거 실패: {e}")
        raise HTTPException(status_code=500, detail=f"집계 트리거 실패: {str(e)}")


@router.get("/report")
async def generate_report(
    date_from: date = Query(..., description="시작 날짜"),
    date_to: date = Query(..., description="종료 날짜"),
    db: Session = Depends(get_db)
):
    """
    종합 리포트 생성

    Args:
        date_from: 시작 날짜
        date_to: 종료 날짜
        db: 데이터베이스 세션

    Returns:
        dict: 종합 리포트
            - period: 조회 기간
            - overview: 전체 개요
            - performance: 성능 메트릭
            - quality: 품질 메트릭
            - usage_patterns: 사용 패턴
            - collections: 컬렉션별 통계
    """
    try:
        if date_to < date_from:
            raise ValueError("종료 날짜는 시작 날짜 이후여야 합니다")

        if (date_to - date_from).days > 90:
            raise ValueError("리포트 기간은 최대 90일까지 가능합니다")

        report = await statistics_service.generate_report(
            date_from=date_from,
            date_to=date_to,
            db=db
        )

        return report

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"리포트 생성 실패: {e}")
        raise HTTPException(status_code=500, detail=f"리포트 생성 실패: {str(e)}")


@router.post("/cleanup")
async def cleanup_old_data(
    background_tasks: BackgroundTasks,
    retention_days: int = Query(30, ge=7, le=365, description="보존 기간 (일)")
):
    """
    오래된 데이터 정리 (관리자용)

    Args:
        retention_days: 보존 기간 (기본: 30일)
        background_tasks: 백그라운드 태스크

    Returns:
        dict: 정리 작업 상태
    """
    try:
        # 백그라운드로 정리 실행
        async def run_cleanup():
            deleted_count = await conversation_service.cleanup_old_conversations(retention_days)
            logger.info(f"정리 완료: {deleted_count}개 파일 삭제")

        background_tasks.add_task(run_cleanup)

        return {
            "message": f"{retention_days}일 이상 된 대화 파일 정리 작업이 시작되었습니다",
            "retention_days": retention_days,
            "status": "started"
        }

    except Exception as e:
        logger.error(f"정리 작업 시작 실패: {e}")
        raise HTTPException(status_code=500, detail=f"정리 작업 시작 실패: {str(e)}")


@router.get("/collections")
async def get_collection_stats(db: Session = Depends(get_db)):
    """
    컬렉션별 통계 개요

    Returns:
        dict: 컬렉션별 통계
    """
    try:
        # 각 컬렉션의 최신 통계 조회
        collections = db.query(ChatStatistics.collection_name).distinct().all()

        collection_stats = []
        for (collection_name,) in collections:
            # 최신 통계 가져오기
            latest_stat = db.query(ChatStatistics).filter(
                ChatStatistics.collection_name == collection_name
            ).order_by(desc(ChatStatistics.date)).first()

            if latest_stat:
                collection_stats.append({
                    "collection_name": collection_name,
                    "latest_date": latest_stat.date.isoformat(),
                    "total_queries": latest_stat.total_queries,
                    "unique_sessions": latest_stat.unique_sessions,
                    "avg_response_time_ms": latest_stat.avg_response_time_ms
                })

        return {
            "collections": collection_stats,
            "total_collections": len(collection_stats)
        }

    except Exception as e:
        logger.error(f"컬렉션 통계 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=f"컬렉션 통계 조회 실패: {str(e)}")


# ============================================================
# 1. 사용자 행동 분석 API
# ============================================================

@router.get("/hourly-heatmap")
async def get_hourly_heatmap(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름 ('ALL'은 전체 조회)"),
    days: int = Query(7, ge=1, le=30, description="최근 N일간 데이터"),
    db: Session = Depends(get_db)
):
    """
    시간대별 사용량 히트맵 데이터

    Returns:
        dict: 요일(0-6) x 시간(0-23) 매트릭스
    """
    try:
        # "ALL" → None 정규화 (전체 조회)
        normalized_collection = _normalize_collection(collection_name)

        end_date = date.today()
        start_date = end_date - timedelta(days=days)

        # 로그 데이터 조회
        df = await statistics_service.query_logs_by_date_range(start_date, end_date, normalized_collection)

        if df.empty:
            # 빈 히트맵 반환
            return {
                "heatmap": [[0]*24 for _ in range(7)],
                "max_value": 0,
                "period": {"from": start_date.isoformat(), "to": end_date.isoformat()}
            }

        # 시간대별 집계
        if 'created_at' in df.columns:
            df['created_at'] = pd.to_datetime(df['created_at'])
            df['hour'] = df['created_at'].dt.hour
            df['dayofweek'] = df['created_at'].dt.dayofweek  # 0=Monday, 6=Sunday

            # 사용자 메시지만 필터링 (쿼리 수만 카운팅, assistant 메시지 제외)
            user_df = df[df['message_type'] == 'user'] if 'message_type' in df.columns else df

            # 히트맵 매트릭스 생성 (7일 x 24시간)
            heatmap = [[0]*24 for _ in range(7)]

            for _, row in user_df.iterrows():
                day = int(row['dayofweek'])
                hour = int(row['hour'])
                heatmap[day][hour] += 1

            max_value = max(max(row) for row in heatmap)

            return {
                "heatmap": heatmap,
                "max_value": max_value,
                "period": {"from": start_date.isoformat(), "to": end_date.isoformat()},
                "labels": {
                    "days": ["월", "화", "수", "목", "금", "토", "일"],
                    "hours": list(range(24))
                }
            }

        return {"heatmap": [[0]*24 for _ in range(7)], "max_value": 0}

    except Exception as e:
        logger.error(f"시간대별 히트맵 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/conversation-stats")
async def get_conversation_stats(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름 ('ALL'은 전체 조회)"),
    days: int = Query(7, ge=1, le=30, description="최근 N일간 데이터"),
    db: Session = Depends(get_db)
):
    """
    대화 통계 (평균 턴수, 재방문율 등)

    Returns:
        dict: 대화 관련 통계
    """
    try:
        # "ALL" → None 정규화 (전체 조회)
        normalized_collection = _normalize_collection(collection_name)

        # ChatSession 테이블에서 조회
        query = db.query(ChatSession)

        end_date = now_naive()
        start_date = end_date - timedelta(days=days)

        query = query.filter(ChatSession.started_at >= start_date)

        if normalized_collection:
            query = query.filter(ChatSession.collection_name == normalized_collection)

        sessions = query.all()

        if not sessions:
            return {
                "avg_turns": 0,
                "avg_user_messages": 0,
                "revisit_rate": 0,
                "total_sessions": 0,
                "regeneration_rate": 0
            }

        # 평균 턴수 계산
        total_turns = sum(s.message_count or 0 for s in sessions)
        avg_turns = total_turns / len(sessions) if sessions else 0

        # 평균 사용자 메시지 수
        total_user_msgs = sum(s.user_message_count or 0 for s in sessions)
        avg_user_messages = total_user_msgs / len(sessions) if sessions else 0

        # 재방문율 (같은 user_hash가 여러 세션을 가진 비율)
        user_hashes = [s.user_hash for s in sessions if s.user_hash]
        if user_hashes:
            from collections import Counter
            hash_counts = Counter(user_hashes)
            revisit_users = sum(1 for count in hash_counts.values() if count > 1)
            revisit_rate = (revisit_users / len(hash_counts)) * 100 if hash_counts else 0
        else:
            revisit_rate = 0

        # 재생성 비율
        regeneration_count = sum(1 for s in sessions if s.has_regeneration)
        regeneration_rate = (regeneration_count / len(sessions)) * 100 if sessions else 0

        return {
            "avg_turns": round(avg_turns, 1),
            "avg_user_messages": round(avg_user_messages, 1),
            "revisit_rate": round(revisit_rate, 1),
            "total_sessions": len(sessions),
            "unique_users": len(set(s.user_hash for s in sessions if s.user_hash)),
            "regeneration_rate": round(regeneration_rate, 1),
            "period": {"from": start_date.isoformat(), "to": end_date.isoformat()}
        }

    except Exception as e:
        logger.error(f"대화 통계 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# 3. 문서/컬렉션 분석 API
# ============================================================

@router.get("/top-documents")
async def get_top_documents(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름 ('ALL'은 전체 조회)"),
    days: int = Query(7, ge=1, le=30, description="최근 N일간 데이터"),
    limit: int = Query(10, ge=1, le=50, description="조회할 문서 수")
):
    """
    가장 많이 참조된 문서 TOP N

    Returns:
        list: 문서별 참조 횟수
    """
    try:
        # "ALL" → None 정규화 (전체 조회)
        normalized_collection = _normalize_collection(collection_name)

        end_date = date.today()
        start_date = end_date - timedelta(days=days)

        df = await statistics_service.query_logs_by_date_range(start_date, end_date, normalized_collection)

        if df.empty or 'retrieval_info' not in df.columns:
            return {"documents": [], "total": 0}

        # retrieval_info에서 문서 정보 추출
        doc_counts = {}

        for info in df['retrieval_info'].dropna():
            if isinstance(info, dict) and 'sources' in info:
                for source in info['sources']:
                    if isinstance(source, dict):
                        doc_name = source.get('document_name') or source.get('source', 'Unknown')
                        doc_counts[doc_name] = doc_counts.get(doc_name, 0) + 1

        # 정렬 및 상위 N개
        sorted_docs = sorted(doc_counts.items(), key=lambda x: x[1], reverse=True)[:limit]

        return {
            "documents": [
                {"name": name, "count": count, "percentage": round(count / sum(doc_counts.values()) * 100, 1)}
                for name, count in sorted_docs
            ],
            "total": len(doc_counts),
            "period": {"from": start_date.isoformat(), "to": end_date.isoformat()}
        }

    except Exception as e:
        logger.error(f"TOP 문서 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/page-distribution")
async def get_page_distribution(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름 ('ALL'은 전체 조회)"),
    document_name: Optional[str] = Query(None, description="특정 문서 필터"),
    days: int = Query(7, ge=1, le=30, description="최근 N일간 데이터")
):
    """
    문서별 페이지 참조 분포

    Returns:
        dict: 페이지별 참조 횟수
    """
    try:
        # "ALL" → None 정규화 (전체 조회)
        normalized_collection = _normalize_collection(collection_name)

        end_date = date.today()
        start_date = end_date - timedelta(days=days)

        df = await statistics_service.query_logs_by_date_range(start_date, end_date, normalized_collection)

        if df.empty or 'retrieval_info' not in df.columns:
            return {"pages": [], "document_stats": {}}

        # 페이지별 참조 횟수
        page_counts = {}
        doc_page_counts = {}  # 문서별 페이지 분포

        for info in df['retrieval_info'].dropna():
            if isinstance(info, dict) and 'sources' in info:
                for source in info['sources']:
                    if isinstance(source, dict):
                        doc_name = source.get('document_name') or source.get('source', 'Unknown')
                        page_num = source.get('page_number') or source.get('page', 0)

                        # 특정 문서 필터
                        if document_name and doc_name != document_name:
                            continue

                        # 전체 페이지 카운트
                        page_key = f"{doc_name}:p{page_num}"
                        page_counts[page_key] = page_counts.get(page_key, 0) + 1

                        # 문서별 페이지 분포
                        if doc_name not in doc_page_counts:
                            doc_page_counts[doc_name] = {}
                        doc_page_counts[doc_name][page_num] = doc_page_counts[doc_name].get(page_num, 0) + 1

        # 상위 페이지 정렬
        sorted_pages = sorted(page_counts.items(), key=lambda x: x[1], reverse=True)[:20]

        return {
            "pages": [
                {"page": page, "count": count}
                for page, count in sorted_pages
            ],
            "document_stats": {
                doc: {
                    "total_refs": sum(pages.values()),
                    "page_count": len(pages),
                    "top_pages": sorted(pages.items(), key=lambda x: x[1], reverse=True)[:5]
                }
                for doc, pages in doc_page_counts.items()
            },
            "period": {"from": start_date.isoformat(), "to": end_date.isoformat()}
        }

    except Exception as e:
        logger.error(f"페이지 분포 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# 4. 성능 모니터링 API
# ============================================================

@router.get("/response-time-distribution")
async def get_response_time_distribution(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름 ('ALL'은 전체 조회)"),
    days: int = Query(7, ge=1, le=30, description="최근 N일간 데이터")
):
    """
    응답시간 분포 히스토그램 데이터

    Returns:
        dict: 응답시간 구간별 빈도 및 백분위
    """
    try:
        # "ALL" → None 정규화 (전체 조회)
        normalized_collection = _normalize_collection(collection_name)

        end_date = date.today()
        start_date = end_date - timedelta(days=days)

        df = await statistics_service.query_logs_by_date_range(start_date, end_date, normalized_collection)

        if df.empty or 'performance' not in df.columns:
            return {
                "histogram": [],
                "percentiles": {"p50": 0, "p90": 0, "p95": 0, "p99": 0},
                "stats": {"min": 0, "max": 0, "avg": 0}
            }

        # 응답시간 추출
        response_times = df['performance'].apply(
            lambda x: x.get('response_time_ms', None) if isinstance(x, dict) else None
        ).dropna().tolist()

        if not response_times:
            return {
                "histogram": [],
                "percentiles": {"p50": 0, "p90": 0, "p95": 0, "p99": 0},
                "stats": {"min": 0, "max": 0, "avg": 0}
            }

        import numpy as np
        response_times = np.array(response_times)

        # 히스토그램 구간 생성 (0-500, 500-1000, 1000-2000, 2000-5000, 5000+)
        bins = [0, 500, 1000, 2000, 5000, 10000, float('inf')]
        bin_labels = ["0-500ms", "500ms-1s", "1-2s", "2-5s", "5-10s", "10s+"]

        histogram = []
        for i in range(len(bins) - 1):
            count = int(np.sum((response_times >= bins[i]) & (response_times < bins[i+1])))
            histogram.append({
                "range": bin_labels[i],
                "count": count,
                "percentage": round(count / len(response_times) * 100, 1)
            })

        # 백분위 계산
        percentiles = {
            "p50": float(np.percentile(response_times, 50)),
            "p90": float(np.percentile(response_times, 90)),
            "p95": float(np.percentile(response_times, 95)),
            "p99": float(np.percentile(response_times, 99))
        }

        return {
            "histogram": histogram,
            "percentiles": percentiles,
            "stats": {
                "min": float(np.min(response_times)),
                "max": float(np.max(response_times)),
                "avg": float(np.mean(response_times)),
                "total_count": len(response_times)
            },
            "period": {"from": start_date.isoformat(), "to": end_date.isoformat()}
        }

    except Exception as e:
        logger.error(f"응답시간 분포 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/token-usage-trend")
async def get_token_usage_trend(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름 ('ALL'은 전체 조회)"),
    days: int = Query(7, ge=1, le=30, description="최근 N일간 데이터")
):
    """
    토큰 사용량 추이

    Returns:
        list: 일별 토큰 사용량
    """
    try:
        # "ALL" → None 정규화 (전체 조회)
        normalized_collection = _normalize_collection(collection_name)

        end_date = date.today()
        start_date = end_date - timedelta(days=days)

        df = await statistics_service.query_logs_by_date_range(start_date, end_date, normalized_collection)

        if df.empty:
            return {"trend": [], "total": 0}

        # 날짜별 토큰 집계
        if 'created_at' in df.columns and 'performance' in df.columns:
            df['date'] = pd.to_datetime(df['created_at']).dt.date
            df['tokens'] = df['performance'].apply(
                lambda x: x.get('token_count', 0) if isinstance(x, dict) else 0
            )

            daily_tokens = df.groupby('date')['tokens'].sum().reset_index()

            trend = [
                {"date": row['date'].isoformat(), "tokens": int(row['tokens'])}
                for _, row in daily_tokens.iterrows()
            ]

            return {
                "trend": sorted(trend, key=lambda x: x['date']),
                "total": int(df['tokens'].sum()),
                "avg_daily": int(df['tokens'].sum() / days) if days > 0 else 0,
                "period": {"from": start_date.isoformat(), "to": end_date.isoformat()}
            }

        return {"trend": [], "total": 0}

    except Exception as e:
        logger.error(f"토큰 사용량 추이 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# 6. 실시간/운영 API
# ============================================================

@router.get("/active-sessions")
async def get_active_sessions(
    minutes: int = Query(5, ge=1, le=60, description="최근 N분 이내 활성"),
    db: Session = Depends(get_db)
):
    """
    실시간 활성 세션 수

    Returns:
        dict: 활성 세션 정보
    """
    try:
        threshold = now_naive() - timedelta(minutes=minutes)

        # 최근 N분 내에 시작된 세션만 활성으로 간주
        active_sessions = db.query(ChatSession).filter(
            ChatSession.started_at >= threshold
        ).all()

        # 컬렉션별 활성 세션 수
        collection_counts = {}
        for session in active_sessions:
            col = session.collection_name
            collection_counts[col] = collection_counts.get(col, 0) + 1

        return {
            "active_count": len(active_sessions),
            "by_collection": collection_counts,
            "threshold_minutes": minutes,
            "timestamp": now_naive().isoformat()
        }

    except Exception as e:
        logger.error(f"활성 세션 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/recent-queries")
async def get_recent_queries(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름"),
    limit: int = Query(20, ge=1, le=100, description="조회할 개수")
):
    """
    최근 질문 실시간 피드

    Returns:
        list: 최근 질문 목록 (최근 7일간)
    """
    try:
        # 최근 7일간 로그에서 최근 질문 조회
        end_date = date.today()
        start_date = end_date - timedelta(days=6)
        df = await statistics_service.query_logs_by_date_range(start_date, end_date, collection_name)

        if df.empty:
            return {"queries": [], "total": 0}

        # 사용자 메시지만 필터링
        if 'message_type' in df.columns:
            user_messages = df[df['message_type'] == 'user'].copy()
        else:
            user_messages = df.copy()

        if user_messages.empty:
            return {"queries": [], "total": 0}

        # 최근 순 정렬
        if 'created_at' in user_messages.columns:
            user_messages['created_at'] = pd.to_datetime(user_messages['created_at'])
            user_messages = user_messages.sort_values('created_at', ascending=False)

        # 상위 N개 추출
        recent = user_messages.head(limit)

        queries = []
        for _, row in recent.iterrows():
            query_data = {
                "query": row.get('message_content', '')[:200],  # 200자 제한
                "collection": row.get('collection_name', ''),
                "timestamp": row['created_at'].isoformat() if 'created_at' in row and pd.notna(row['created_at']) else None,
                "session_id": row.get('session_id', '')[:8] if row.get('session_id') else None  # 앞 8자리만
            }

            # 성능 정보 추가
            if 'performance' in row and isinstance(row['performance'], dict):
                query_data['response_time_ms'] = row['performance'].get('response_time_ms')

            queries.append(query_data)

        return {
            "queries": queries,
            "total": len(user_messages),
            "timestamp": now_naive().isoformat()
        }

    except Exception as e:
        logger.error(f"최근 질문 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# 7. Excel 내보내기 API
# ============================================================

@router.get("/export/excel")
async def export_conversations_to_excel(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름 (미지정 또는 ALL시 전체)"),
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    db: Session = Depends(get_db)
):
    """
    대화 내역을 Excel 파일로 내보내기

    Args:
        collection_name: 컬렉션 이름 (선택, 미지정 또는 ALL시 전체)
        date_from: 시작 날짜 (선택, 기본값: 오늘)
        date_to: 종료 날짜 (선택, 기본값: 오늘)
        db: 데이터베이스 세션

    Returns:
        StreamingResponse: Excel 파일 스트림
    """
    try:
        # 날짜 기본값 설정
        if not date_from:
            date_from = date.today()
        if not date_to:
            date_to = date.today()

        # "ALL"이면 None으로 변환 (전체 조회)
        effective_collection = None if collection_name in (None, "ALL") else collection_name

        # 날짜 변환
        start_datetime = datetime.combine(date_from, datetime.min.time())
        end_datetime = datetime.combine(date_to, datetime.max.time())

        # conversations 로그에서 데이터 조회
        conversations = await conversation_service.read_conversations(
            start_date=start_datetime,
            end_date=end_datetime,
            collection_name=effective_collection,
            limit=10000  # 최대 10000건
        )

        # data 로그에서 추가 메타데이터 조회
        logs_df = await statistics_service.query_logs_by_date_range(
            date_from, date_to, effective_collection
        )

        # Q&A 쌍으로 데이터 가공
        export_data = []

        for conv in conversations:
            conv_id = conv.get("conversation_id", "")
            conv_collection = conv.get("collection_name", "")
            metadata = conv.get("metadata", {})
            started_at = conv.get("started_at", "")
            messages = conv.get("messages", [])

            # 메시지를 Q&A 쌍으로 그룹화
            i = 0
            while i < len(messages):
                user_msg = None
                assistant_msg = None
                retrieved_docs = []

                # user 메시지 찾기
                if i < len(messages) and messages[i].get("role") == "user":
                    user_msg = messages[i]
                    i += 1

                # assistant 메시지 찾기
                if i < len(messages) and messages[i].get("role") == "assistant":
                    assistant_msg = messages[i]
                    retrieved_docs = assistant_msg.get("retrieved_docs", [])
                    i += 1

                if user_msg:
                    # 날짜/시간 형식 변환 (ISO -> yyyy-mm-dd hh:mm)
                    timestamp_raw = user_msg.get("timestamp", started_at)
                    try:
                        if isinstance(timestamp_raw, str) and timestamp_raw:
                            # ISO 형식 파싱 후 포맷 변환
                            dt = datetime.fromisoformat(timestamp_raw.replace("Z", "+00:00"))
                            formatted_timestamp = dt.strftime("%Y-%m-%d %H:%M")
                        else:
                            formatted_timestamp = str(timestamp_raw)
                    except (ValueError, TypeError):
                        formatted_timestamp = str(timestamp_raw)

                    # 참조 문서 정보 추출 (챗봇 스타일)
                    sources_list = []
                    scores_list = []
                    for idx, doc in enumerate(retrieved_docs, 1):
                        doc_metadata = doc.get("metadata", {})
                        source_name = doc_metadata.get("source_file", doc_metadata.get("source", "Unknown"))
                        page_num = doc_metadata.get("page_number", doc_metadata.get("page", "-"))
                        section = doc_metadata.get("section", doc_metadata.get("headings", ""))
                        score = doc.get("score", 0)
                        score_pct = round(score * 100, 1)

                        # 챗봇 스타일 형식: #1. [99.9%] 파일명 (p.30) - 섹션
                        source_str = f"#{idx}. [{score_pct}%] {source_name}"
                        if page_num and page_num != "-":
                            source_str += f" (p.{page_num})"
                        if section:
                            # 섹션이 리스트인 경우 처리
                            if isinstance(section, list):
                                section = " > ".join(str(s) for s in section[:2])  # 최대 2개
                            source_str += f" - {section[:50]}"  # 최대 50자

                        sources_list.append(source_str)
                        scores_list.append(score)

                    # data 로그에서 추가 정보 찾기
                    performance_info = {}
                    llm_info = {}
                    if not logs_df.empty and 'session_id' in logs_df.columns:
                        # session_id로 매칭 시도
                        matching_logs = logs_df[
                            (logs_df.get('message_type') == 'assistant') &
                            (logs_df.get('message_content', '').str[:50] == (assistant_msg.get("content", "")[:50] if assistant_msg else ""))
                        ] if assistant_msg else pd.DataFrame()

                        if not matching_logs.empty:
                            first_match = matching_logs.iloc[0]
                            if 'performance' in first_match and isinstance(first_match['performance'], dict):
                                performance_info = first_match['performance']
                            if 'llm_model' in first_match:
                                llm_info['model'] = first_match['llm_model']
                            if 'reasoning_level' in first_match:
                                llm_info['reasoning_level'] = first_match['reasoning_level']

                    row = {
                        "날짜/시간": formatted_timestamp,
                        "세션 ID": conv_id[:8] if conv_id else "",
                        "사용자 질문": user_msg.get("content", ""),
                        "AI 응답": assistant_msg.get("content", "") if assistant_msg else "",
                        "컬렉션": conv_collection,
                        "응답시간(ms)": performance_info.get("response_time_ms", metadata.get("duration_seconds", 0) * 1000 if metadata.get("duration_seconds") else ""),
                        "토큰수": performance_info.get("token_count", ""),
                        "검색점수": round(scores_list[0], 4) if scores_list else "",
                        "참조문서": "\n".join(sources_list) if sources_list else "",
                        "에러여부": "Y" if metadata.get("has_error") else "N",
                        "재생성여부": "Y" if metadata.get("has_regeneration") else "N",
                        "LLM모델": llm_info.get("model", ""),
                        "추론레벨": llm_info.get("reasoning_level", ""),
                    }
                    export_data.append(row)

        # 헤더 정의
        headers = [
            "날짜/시간", "세션 ID", "사용자 질문", "AI 응답", "컬렉션",
            "응답시간(ms)", "토큰수", "검색점수", "참조문서",
            "에러여부", "재생성여부", "LLM모델", "추론레벨"
        ]

        # Excel 워크북 생성 (write_only=True로 메모리 최적화)
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="대화내역")

        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)

        # 헤더 행 작성 (스타일 적용)
        header_row = []
        for header in headers:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            header_row.append(cell)
        ws.append(header_row)

        # 데이터 작성 (각 행을 순차적으로 쓰고 메모리 해제)
        for row_data in export_data:
            data_row = []
            for header in headers:
                value = row_data.get(header, "")
                cell = WriteOnlyCell(ws, value=value)
                cell.alignment = cell_alignment
                data_row.append(cell)
            ws.append(data_row)

        # write_only 모드에서는 column_dimensions와 freeze_panes 설정 불가
        # 대신 클라이언트 측에서 처리하거나 별도 후처리 필요

        # 메모리 스트림에 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 파일명 생성 (한글 포함 시 URL 인코딩)
        filename = f"conversations_{collection_name}_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
        encoded_filename = quote(filename, safe='')

        # StreamingResponse로 반환
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        logger.error(f"Excel 내보내기 실패: {e}")
        raise HTTPException(status_code=500, detail=f"Excel 내보내기 실패: {str(e)}")


# ============================================================
# 8. 오류 로그 다운로드 API
# ============================================================

@router.get("/errors/download")
async def download_error_logs(
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    db: Session = Depends(get_db)
):
    """
    오류 로그를 Excel 파일로 다운로드

    Args:
        date_from: 시작 날짜 (선택, 기본값: 7일 전)
        date_to: 종료 날짜 (선택, 기본값: 오늘)
        db: 데이터베이스 세션

    Returns:
        StreamingResponse: Excel 파일 스트림
    """
    try:
        # 날짜 기본값 설정
        if not date_to:
            date_to = date.today()
        if not date_from:
            date_from = date_to - timedelta(days=7)

        # 로그 데이터 조회
        df = await statistics_service.query_logs_by_date_range(date_from, date_to, None)

        if df.empty:
            # 빈 Excel 반환
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="오류 로그")
            ws.append(["오류 데이터가 없습니다."])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"error_logs_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
            encoded_filename = quote(filename, safe='')

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )

        # error_info가 있는 항목만 필터링
        error_logs = []

        for _, row in df.iterrows():
            error_info = row.get('error_info')
            if error_info and isinstance(error_info, dict) and error_info:
                # 사용자 메시지만 (assistant 응답이 아닌)
                if row.get('message_type') == 'user':
                    # 날짜/시간 형식 변환
                    timestamp_raw = row.get('created_at', '')
                    try:
                        if isinstance(timestamp_raw, str) and timestamp_raw:
                            dt = datetime.fromisoformat(timestamp_raw.replace("Z", "+00:00"))
                            formatted_timestamp = dt.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            formatted_timestamp = str(timestamp_raw)
                    except (ValueError, TypeError):
                        formatted_timestamp = str(timestamp_raw)

                    # 클라이언트 정보
                    client_info = row.get('client_info', {}) or {}

                    error_logs.append({
                        "발생일시": formatted_timestamp,
                        "세션ID": str(row.get('session_id', ''))[:12],
                        "컬렉션": row.get('collection_name', ''),
                        "사용자 질문": str(row.get('message_content', ''))[:500],
                        "모델": row.get('llm_model', ''),
                        "오류유형": error_info.get('type', error_info.get('error_type', 'Unknown')),
                        "오류메시지": str(error_info.get('message', error_info.get('error_message', str(error_info))))[:1000],
                        "IP해시": client_info.get('ip_hash', '')[:16] if client_info.get('ip_hash') else '',
                    })

        # Excel 워크북 생성
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="오류 로그")

        # 헤더 정의
        headers = ["발생일시", "세션ID", "컬렉션", "사용자 질문", "모델", "오류유형", "오류메시지", "IP해시"]

        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")  # 빨간색
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)

        # 헤더 행 작성
        header_row = []
        for header in headers:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            header_row.append(cell)
        ws.append(header_row)

        # 데이터 작성
        for row_data in error_logs:
            data_row = []
            for header in headers:
                value = row_data.get(header, "")
                cell = WriteOnlyCell(ws, value=value)
                cell.alignment = cell_alignment
                data_row.append(cell)
            ws.append(data_row)

        # 요약 시트 추가
        summary_ws = wb.create_sheet(title="요약")
        summary_ws.append(["항목", "값"])
        summary_ws.append(["조회 기간", f"{date_from.isoformat()} ~ {date_to.isoformat()}"])
        summary_ws.append(["총 오류 건수", len(error_logs)])
        summary_ws.append(["생성 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        # 메모리 스트림에 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 파일명 생성
        filename = f"error_logs_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
        encoded_filename = quote(filename, safe='')

        logger.info(f"오류 로그 Excel 다운로드: {len(error_logs)}건, 기간: {date_from} ~ {date_to}")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except Exception as e:
        logger.error(f"오류 로그 다운로드 실패: {e}")
        raise HTTPException(status_code=500, detail=f"오류 로그 다운로드 실패: {str(e)}")