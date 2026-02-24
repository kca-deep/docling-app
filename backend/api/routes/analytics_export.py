"""
Analytics Excel 내보내기 API 라우터
대화 내역 및 오류 로그 Excel 다운로드 엔드포인트
"""
import io
import logging
from datetime import datetime, date, timedelta
from typing import Optional
from urllib.parse import quote

import pandas as pd
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.cell import WriteOnlyCell
from sqlalchemy.orm import Session

from backend.database import get_db
from backend.dependencies.auth import get_current_active_user
from backend.utils.timezone import now_naive
from backend.services.statistics_service import statistics_service
from backend.services.conversation_service import conversation_service

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/analytics",
    tags=["analytics-export"],
    dependencies=[Depends(get_current_active_user)]
)


def _write_styled_excel_sheet(
    wb: Workbook,
    title: str,
    headers: list,
    data: list,
    header_color: str = "4472C4"
):
    """
    스타일이 적용된 Excel 시트 생성 (write_only 모드)

    Args:
        wb: write_only Workbook
        title: 시트 제목
        headers: 헤더 목록
        data: 딕셔너리 리스트 (각 dict의 키가 headers와 매핑)
        header_color: 헤더 배경색 (기본: 파란색)

    Returns:
        워크시트 객체
    """
    ws = wb.create_sheet(title=title)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    header_row = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        header_row.append(cell)
    ws.append(header_row)

    for row_data in data:
        data_row = []
        for header in headers:
            value = row_data.get(header, "")
            cell = WriteOnlyCell(ws, value=value)
            cell.alignment = cell_alignment
            data_row.append(cell)
        ws.append(data_row)

    return ws


@router.get("/export/excel")
async def export_conversations_to_excel(
    collection_name: Optional[str] = Query(None, description="컬렉션 이름 (미지정 또는 ALL시 전체)"),
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    db: Session = Depends(get_db)
):
    """
    대화 내역을 Excel 파일로 내보내기

    Args:
        collection_name: 컬렉션 이름 (선택, 미지정 또는 ALL시 전체)
        date_from: 시작 날짜 (선택, 기본값: 오늘)
        date_to: 종료 날짜 (선택, 기본값: 오늘)
        db: 데이터베이스 세션

    Returns:
        StreamingResponse: Excel 파일 스트림
    """
    try:
        if not date_from:
            date_from = date.today()
        if not date_to:
            date_to = date.today()

        effective_collection = None if collection_name in (None, "ALL") else collection_name

        start_datetime = datetime.combine(date_from, datetime.min.time())
        end_datetime = datetime.combine(date_to, datetime.max.time())

        conversations = await conversation_service.read_conversations(
            start_date=start_datetime,
            end_date=end_datetime,
            collection_name=effective_collection,
            limit=10000
        )

        logs_df = await statistics_service.query_logs_by_date_range(
            date_from, date_to, effective_collection
        )

        export_data = []

        for conv in conversations:
            conv_id = conv.get("conversation_id", "")
            conv_collection = conv.get("collection_name", "")
            metadata = conv.get("metadata", {})
            started_at = conv.get("started_at", "")
            messages = conv.get("messages", [])

            i = 0
            while i < len(messages):
                user_msg = None
                assistant_msg = None
                retrieved_docs = []

                if i < len(messages) and messages[i].get("role") == "user":
                    user_msg = messages[i]
                    i += 1

                if i < len(messages) and messages[i].get("role") == "assistant":
                    assistant_msg = messages[i]
                    retrieved_docs = assistant_msg.get("retrieved_docs", [])
                    i += 1

                if user_msg:
                    timestamp_raw = user_msg.get("timestamp", started_at)
                    try:
                        if isinstance(timestamp_raw, str) and timestamp_raw:
                            dt = datetime.fromisoformat(timestamp_raw.replace("Z", "+00:00"))
                            formatted_timestamp = dt.strftime("%Y-%m-%d %H:%M")
                        else:
                            formatted_timestamp = str(timestamp_raw)
                    except (ValueError, TypeError):
                        formatted_timestamp = str(timestamp_raw)

                    sources_list = []
                    scores_list = []
                    for idx, doc in enumerate(retrieved_docs, 1):
                        doc_metadata = doc.get("metadata", {})
                        source_name = doc_metadata.get("filename") or doc_metadata.get("source_file", doc_metadata.get("source", "Unknown"))
                        page_num = doc_metadata.get("page_number", doc_metadata.get("page", "-"))
                        section = doc_metadata.get("section", doc_metadata.get("headings", ""))
                        score = doc.get("score", 0)
                        score_pct = round(score * 100, 1)

                        source_str = f"#{idx}. [{score_pct}%] {source_name}"
                        if page_num and page_num != "-":
                            source_str += f" (p.{page_num})"
                        if section:
                            if isinstance(section, list):
                                section = " > ".join(str(s) for s in section[:2])
                            source_str += f" - {section[:50]}"

                        sources_list.append(source_str)
                        scores_list.append(score)

                    performance_info = {}
                    llm_info = {}
                    if not logs_df.empty and 'session_id' in logs_df.columns:
                        matching_logs = logs_df[
                            (logs_df.get('message_type') == 'assistant') &
                            (logs_df.get('message_content', '').str[:50] == (assistant_msg.get("content", "")[:50] if assistant_msg else ""))
                        ] if assistant_msg else pd.DataFrame()

                        if not matching_logs.empty:
                            first_match = matching_logs.iloc[0]
                            if 'performance' in first_match and isinstance(first_match['performance'], dict):
                                performance_info = first_match['performance']
                            if 'llm_model' in first_match:
                                llm_info['model'] = first_match['llm_model']
                            if 'reasoning_level' in first_match:
                                llm_info['reasoning_level'] = first_match['reasoning_level']

                    client_ip = ""
                    client_ip_hash = ""
                    if not logs_df.empty and 'client_info' in logs_df.columns:
                        session_logs = logs_df[logs_df.get('session_id') == conv_id] if 'session_id' in logs_df.columns else pd.DataFrame()
                        if not session_logs.empty:
                            first_log = session_logs.iloc[0]
                            if 'client_info' in first_log and isinstance(first_log['client_info'], dict):
                                client_ip = first_log['client_info'].get('ip', '')
                                client_ip_hash = first_log['client_info'].get('ip_hash', '')

                    row = {
                        "날짜/시간": formatted_timestamp,
                        "세션 ID": conv_id[:8] if conv_id else "",
                        "사용자 질문": user_msg.get("content", ""),
                        "AI 응답": assistant_msg.get("content", "") if assistant_msg else "",
                        "컬렉션": conv_collection,
                        "응답시간(ms)": performance_info.get("response_time_ms", metadata.get("duration_seconds", 0) * 1000 if metadata.get("duration_seconds") else ""),
                        "토큰수": performance_info.get("token_count", ""),
                        "검색점수": round(scores_list[0], 4) if scores_list else "",
                        "참조문서": "\n".join(sources_list) if sources_list else "",
                        "에러여부": "Y" if metadata.get("has_error") else "N",
                        "재생성여부": "Y" if metadata.get("has_regeneration") else "N",
                        "LLM모델": llm_info.get("model", ""),
                        "추론레벨": llm_info.get("reasoning_level", ""),
                        "IP": client_ip,
                        "IP해시": client_ip_hash[:16] if client_ip_hash else "",
                    }
                    export_data.append(row)

        headers = [
            "날짜/시간", "세션 ID", "사용자 질문", "AI 응답", "컬렉션",
            "응답시간(ms)", "토큰수", "검색점수", "참조문서",
            "에러여부", "재생성여부", "LLM모델", "추론레벨", "IP", "IP해시"
        ]

        wb = Workbook(write_only=True)
        _write_styled_excel_sheet(wb, "대화내역", headers, export_data, header_color="4472C4")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"conversations_{collection_name}_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
        encoded_filename = quote(filename, safe='')

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        logger.error(f"Excel 내보내기 실패: {e}")
        raise HTTPException(status_code=500, detail=f"Excel 내보내기 실패: {str(e)}")


@router.get("/errors/download")
async def download_error_logs(
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    db: Session = Depends(get_db)
):
    """
    오류 로그를 Excel 파일로 다운로드

    Args:
        date_from: 시작 날짜 (선택, 기본값: 7일 전)
        date_to: 종료 날짜 (선택, 기본값: 오늘)
        db: 데이터베이스 세션

    Returns:
        StreamingResponse: Excel 파일 스트림
    """
    try:
        if not date_to:
            date_to = date.today()
        if not date_from:
            date_from = date_to - timedelta(days=7)

        df = await statistics_service.query_logs_by_date_range(date_from, date_to, None)

        if df.empty:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="오류 로그")
            ws.append(["오류 데이터가 없습니다."])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"error_logs_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
            encoded_filename = quote(filename, safe='')

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )

        error_logs = []

        for _, row in df.iterrows():
            error_info = row.get('error_info')
            if error_info and isinstance(error_info, dict) and error_info:
                if row.get('message_type') == 'user':
                    timestamp_raw = row.get('created_at', '')
                    try:
                        if isinstance(timestamp_raw, str) and timestamp_raw:
                            dt = datetime.fromisoformat(timestamp_raw.replace("Z", "+00:00"))
                            formatted_timestamp = dt.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            formatted_timestamp = str(timestamp_raw)
                    except (ValueError, TypeError):
                        formatted_timestamp = str(timestamp_raw)

                    client_info = row.get('client_info', {}) or {}

                    error_logs.append({
                        "발생일시": formatted_timestamp,
                        "세션ID": str(row.get('session_id', ''))[:12],
                        "컬렉션": row.get('collection_name', ''),
                        "사용자 질문": str(row.get('message_content', ''))[:500],
                        "모델": row.get('llm_model', ''),
                        "오류유형": error_info.get('type', error_info.get('error_type', 'Unknown')),
                        "오류메시지": str(error_info.get('message', error_info.get('error_message', str(error_info))))[:1000],
                        "IP": client_info.get('ip', ''),
                        "IP해시": client_info.get('ip_hash', '')[:16] if client_info.get('ip_hash') else '',
                    })

        wb = Workbook(write_only=True)
        headers = ["발생일시", "세션ID", "컬렉션", "사용자 질문", "모델", "오류유형", "오류메시지", "IP", "IP해시"]
        _write_styled_excel_sheet(wb, "오류 로그", headers, error_logs, header_color="C0504D")

        summary_ws = wb.create_sheet(title="요약")
        summary_ws.append(["항목", "값"])
        summary_ws.append(["조회 기간", f"{date_from.isoformat()} ~ {date_to.isoformat()}"])
        summary_ws.append(["총 오류 건수", len(error_logs)])
        summary_ws.append(["생성 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"error_logs_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
        encoded_filename = quote(filename, safe='')

        logger.info(f"오류 로그 Excel 다운로드: {len(error_logs)}건, 기간: {date_from} ~ {date_to}")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except Exception as e:
        logger.error(f"오류 로그 다운로드 실패: {e}")
        raise HTTPException(status_code=500, detail=f"오류 로그 다운로드 실패: {str(e)}")
