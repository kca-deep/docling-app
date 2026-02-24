"""
Qdrant Excel 임베딩 API 라우트
Q&A Excel, 동적 Excel 임베딩, 마이그레이션 엔드포인트
"""
import uuid
import io
import json
import logging
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from sqlalchemy.orm import Session

from backend.services.qdrant_service import qdrant_service
from backend.services.embedding_service import embedding_service

from .excel_utils import detect_column_mapping
from backend.database import get_db
from backend.config.settings import settings
from backend.dependencies.auth import get_current_active_user
from backend.models.user import User
from backend.models.document import Document
from backend.models.qdrant_upload_history import QdrantUploadHistory
from backend.models.schemas import (
    QAPreviewResponse,
    QAPreviewRow,
    QAEmbeddingRequest,
    QAEmbeddingResponse,
    QAEmbeddingResult,
    ExcelPreviewResponse,
    ExcelPreviewRow,
    DynamicEmbeddingRequest,
    DynamicEmbeddingResponse,
    DynamicEmbeddingResult,
)

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/qdrant",
    tags=["qdrant-excel"]
)


# ==================== Q&A Excel Embedding Endpoints ====================

@router.post("/qa/preview", response_model=QAPreviewResponse)
async def preview_qa_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user)
):
    """
    Q&A Excel 파일 미리보기 API

    Args:
        file: 업로드된 Excel 파일 (.xlsx)

    Returns:
        QAPreviewResponse: 미리보기 데이터
    """
    try:
        import openpyxl

        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="Excel 파일(.xlsx, .xls)만 지원합니다"
            )

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        headers = [cell.value for cell in ws[1] if cell.value]
        print(f"[INFO] Excel headers: {headers}")

        required_headers = ['question', 'answer_text']
        header_lower = [h.lower() if h else '' for h in headers]

        for req in required_headers:
            if req not in header_lower:
                raise HTTPException(
                    status_code=400,
                    detail=f"필수 컬럼 '{req}'이(가) 없습니다. 현재 컬럼: {headers}"
                )

        header_map = {h.lower(): i for i, h in enumerate(headers) if h}

        rows = []
        for row_num in range(2, ws.max_row + 1):
            row_values = [ws.cell(row=row_num, column=i+1).value for i in range(len(headers))]

            question = row_values[header_map.get('question', 1)] or ''
            answer_text = row_values[header_map.get('answer_text', 2)] or ''

            if not question.strip() and not answer_text.strip():
                continue

            faq_id = row_values[header_map.get('faq_id', 0)] if 'faq_id' in header_map else f"FAQ-{row_num-1:04d}"
            tags_raw = row_values[header_map.get('tag', -1)] if 'tag' in header_map else ''
            tags = [t.strip() for t in (tags_raw or '').split(',')] if tags_raw else []
            policy_anchor = row_values[header_map.get('policy_anchor', -1)] if 'policy_anchor' in header_map else None
            source = row_values[header_map.get('source', -1)] if 'source' in header_map else None

            rows.append(QAPreviewRow(
                row_index=row_num - 2,
                faq_id=str(faq_id) if faq_id else f"FAQ-{row_num-1:04d}",
                question=str(question),
                answer_text=str(answer_text),
                tags=tags,
                policy_anchor=str(policy_anchor) if policy_anchor else None,
                source=str(source) if source else None
            ))

        wb.close()

        print(f"[INFO] Parsed {len(rows)} Q&A rows from Excel")

        return QAPreviewResponse(
            total_rows=len(rows),
            headers=headers,
            preview_rows=rows,
            file_name=file.filename
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Failed to parse Excel file: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Excel 파일 파싱 실패: {str(e)}"
        )


@router.post("/qa/embed", response_model=QAEmbeddingResponse)
async def embed_qa_rows(
    request: QAEmbeddingRequest,
    current_user: User = Depends(get_current_active_user)
):
    """
    Q&A 행별 임베딩 및 Qdrant 업로드 API

    Args:
        request: Q&A 임베딩 요청

    Returns:
        QAEmbeddingResponse: 임베딩 결과
    """
    results = []
    success_count = 0
    failure_count = 0

    try:
        collection_exists = await qdrant_service.collection_exists(request.collection_name)
        if not collection_exists:
            raise HTTPException(
                status_code=404,
                detail=f"Collection '{request.collection_name}'이 존재하지 않습니다"
            )

        batch_size = settings.UPLOAD_BATCH_SIZE
        rows = request.rows

        for batch_start in range(0, len(rows), batch_size):
            batch_end = min(batch_start + batch_size, len(rows))
            batch_rows = rows[batch_start:batch_end]

            try:
                texts = [f"질문: {row.question}\n답변: {row.answer_text}" for row in batch_rows]

                embeddings = await embedding_service.get_embeddings(texts)

                metadata_list = []
                for row in batch_rows:
                    metadata_list.append({
                        "faq_id": row.faq_id,
                        "question": row.question,
                        "answer_text": row.answer_text,
                        "tags": row.tags,
                        "policy_anchor": row.policy_anchor or "",
                        "source": row.source or "",
                        "row_index": row.row_index
                    })

                vector_ids = await qdrant_service.upsert_vectors(
                    collection_name=request.collection_name,
                    vectors=embeddings,
                    texts=texts,
                    metadata_list=metadata_list
                )

                for i, row in enumerate(batch_rows):
                    results.append(QAEmbeddingResult(
                        row_index=row.row_index,
                        faq_id=row.faq_id,
                        success=True,
                        vector_id=vector_ids[i] if i < len(vector_ids) else None
                    ))
                    success_count += 1

                print(f"[INFO] Embedded batch {batch_start+1}-{batch_end}")

            except Exception as e:
                print(f"[ERROR] Failed to embed batch {batch_start+1}-{batch_end}: {e}")
                for row in batch_rows:
                    results.append(QAEmbeddingResult(
                        row_index=row.row_index,
                        faq_id=row.faq_id,
                        success=False,
                        error=str(e)
                    ))
                    failure_count += 1

        return QAEmbeddingResponse(
            total=len(rows),
            success_count=success_count,
            failure_count=failure_count,
            results=results
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Q&A embedding failed: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Q&A 임베딩 실패: {str(e)}"
        )


# ==================== Dynamic Excel Embedding Endpoints ====================
# 컬럼 감지 로직은 excel_utils.py로 분리됨


@router.post("/excel/preview", response_model=ExcelPreviewResponse)
async def preview_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user)
):
    """
    Excel 파일 미리보기 및 스마트 컬럼 감지 API

    Args:
        file: 업로드된 Excel 파일 (.xlsx)

    Returns:
        ExcelPreviewResponse: 미리보기 데이터 및 감지된 매핑
    """
    try:
        import openpyxl

        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="Excel 파일(.xlsx, .xls)만 지원합니다"
            )

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        headers = [cell.value for cell in ws[1] if cell.value]
        print(f"[INFO] Excel headers: {headers}")

        # 스마트 컬럼 매핑 감지
        detected_mapping = detect_column_mapping(headers)
        print(f"[INFO] Detected mapping: {detected_mapping}")

        # 모든 행 읽기
        rows = []
        for row_num in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False

            for col_num, header in enumerate(headers):
                cell_value = ws.cell(row=row_num, column=col_num + 1).value
                if cell_value is not None:
                    row_data[header] = str(cell_value) if cell_value else ""
                    if str(cell_value).strip():
                        has_content = True
                else:
                    row_data[header] = ""

            if has_content:
                rows.append(ExcelPreviewRow(
                    row_index=row_num - 2,
                    data=row_data
                ))

        wb.close()

        print(f"[INFO] Parsed {len(rows)} rows from Excel")

        return ExcelPreviewResponse(
            total_rows=len(rows),
            headers=headers,
            preview_rows=rows,
            file_name=file.filename,
            detected_mapping=detected_mapping
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Failed to parse Excel file: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Excel 파일 파싱 실패: {str(e)}"
        )


@router.post("/excel/embed", response_model=DynamicEmbeddingResponse)
async def embed_excel_dynamic(
    request: DynamicEmbeddingRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    동적 컬럼 매핑을 사용한 Excel 임베딩 API

    Args:
        request: 동적 임베딩 요청 (컬럼 매핑 포함)

    Returns:
        DynamicEmbeddingResponse: 임베딩 결과
    """
    from backend.services.excel_document_service import (
        generate_excel_metadata_content,
        generate_preview,
        calculate_total_length
    )

    results = []
    success_count = 0
    failure_count = 0
    vector_ids_all = []  # SQLite 기록용 전체 vector ID
    texts_all = []  # 미리보기용 텍스트
    excel_doc = None  # SQLite Document 참조

    try:
        # Collection 존재 확인
        collection_exists = await qdrant_service.collection_exists(request.collection_name)
        if not collection_exists:
            raise HTTPException(
                status_code=404,
                detail=f"Collection '{request.collection_name}'이 존재하지 않습니다"
            )

        mapping = request.mapping
        batch_size = settings.UPLOAD_BATCH_SIZE
        rows = request.rows

        # SQLite Document 먼저 생성 (document_id 획득을 위해)
        rows_dict = [{"row_index": r.row_index, "data": r.data} for r in request.rows]
        mapping_dict = request.mapping.model_dump()

        excel_doc = Document(
            task_id=f"excel-{uuid.uuid4().hex[:12]}",
            original_filename=request.file_name,
            file_size=None,
            file_type="xlsx",
            status="pending",  # 업로드 완료 후 success로 변경
            content_length=calculate_total_length(rows_dict, mapping_dict.get("text_columns", [])),
            content_preview="",  # 업로드 완료 후 업데이트
            md_content=generate_excel_metadata_content(
                request.file_name,
                rows_dict,
                mapping_dict,
                len(request.rows)
            ),
            category=request.collection_name,
            parse_options={
                "source_type": "excel",
                "mapping": mapping_dict,
                "total_rows": len(request.rows)
            }
        )
        db.add(excel_doc)
        db.flush()  # document_id 획득

        logger.info(f"Created Excel document in SQLite: id={excel_doc.id}, filename={request.file_name}")

        for batch_start in range(0, len(rows), batch_size):
            batch_end = min(batch_start + batch_size, len(rows))
            batch_rows = rows[batch_start:batch_end]

            try:
                # 임베딩 텍스트 생성
                texts = []
                for row in batch_rows:
                    if mapping.text_template:
                        # 템플릿 사용
                        text = mapping.text_template
                        for key, value in row.data.items():
                            text = text.replace(f"{{{key}}}", str(value) if value else "")
                    else:
                        # 텍스트 컬럼 연결
                        text_parts = []
                        for col in mapping.text_columns:
                            if col in row.data and row.data[col]:
                                text_parts.append(str(row.data[col]))
                        text = "\n".join(text_parts)

                    texts.append(text)

                # 임베딩 생성
                embeddings = await embedding_service.get_embeddings(texts)

                # 메타데이터 생성 (document_id 포함)
                metadata_list = []
                for row in batch_rows:
                    metadata = {
                        "document_id": excel_doc.id,  # document_id 추가
                        "filename": request.file_name,
                        "row_index": row.row_index
                    }

                    # ID 컬럼
                    if mapping.id_column and mapping.id_column in row.data:
                        metadata["id"] = row.data[mapping.id_column]

                    # 태그 컬럼
                    if mapping.tag_column and mapping.tag_column in row.data:
                        tag_value = row.data[mapping.tag_column]
                        if tag_value:
                            metadata["tags"] = [t.strip() for t in str(tag_value).split(',')]

                    # 메타데이터 컬럼들
                    for col in mapping.metadata_columns:
                        if col in row.data:
                            metadata[col] = row.data[col]

                    # 텍스트 컬럼들도 메타데이터에 저장
                    for col in mapping.text_columns:
                        if col in row.data:
                            metadata[col] = row.data[col]

                    # headings 생성 (참조문서 표시용)
                    if mapping.heading_columns:
                        # 사용자가 지정한 컬럼들로 headings 생성
                        headings = []
                        # 무의미한 값 목록 (참조문서 제목으로 부적합)
                        invalid_heading_values = {'-', '--', '없음', 'N/A', 'n/a', 'NA', 'null', 'None', '해당없음', '해당 없음'}
                        for col in mapping.heading_columns:
                            if col in row.data and row.data[col]:
                                val = str(row.data[col]).strip()
                                # 무의미한 값은 제외하고 유효한 값만 추가
                                if val and val not in invalid_heading_values:
                                    headings.append(val)
                        metadata["headings"] = headings if headings else [request.file_name, f"행 {row.row_index + 1}"]
                    else:
                        # 기본값: [파일명, 행 번호]
                        metadata["headings"] = [request.file_name, f"행 {row.row_index + 1}"]

                    metadata_list.append(metadata)

                # Qdrant 업로드
                vector_ids = await qdrant_service.upsert_vectors(
                    collection_name=request.collection_name,
                    vectors=embeddings,
                    texts=texts,
                    metadata_list=metadata_list
                )

                # 결과 기록
                for i, row in enumerate(batch_rows):
                    id_value = None
                    if mapping.id_column and mapping.id_column in row.data:
                        id_value = row.data[mapping.id_column]

                    results.append(DynamicEmbeddingResult(
                        row_index=row.row_index,
                        id_value=id_value,
                        success=True,
                        vector_id=vector_ids[i] if i < len(vector_ids) else None
                    ))
                    success_count += 1

                print(f"[INFO] Embedded batch {batch_start+1}-{batch_end}")

                # SQLite 기록용 데이터 수집
                vector_ids_all.extend(vector_ids)
                texts_all.extend(texts)

            except Exception as e:
                print(f"[ERROR] Failed to embed Excel batch {batch_start+1}-{batch_end}: {e}")
                for row in batch_rows:
                    id_value = None
                    if mapping.id_column and mapping.id_column in row.data:
                        id_value = row.data[mapping.id_column]

                    results.append(DynamicEmbeddingResult(
                        row_index=row.row_index,
                        id_value=id_value,
                        success=False,
                        error=str(e)
                    ))
                    failure_count += 1

        # SQLite Document 업데이트 및 History 생성
        if success_count > 0:
            try:
                # Document 상태 및 미리보기 업데이트
                excel_doc.status = "success"
                excel_doc.content_preview = generate_preview(texts_all[:5])

                # QdrantUploadHistory 생성
                history = QdrantUploadHistory(
                    document_id=excel_doc.id,
                    collection_name=request.collection_name,
                    chunk_count=success_count,
                    vector_ids_json=json.dumps(vector_ids_all),
                    qdrant_url=settings.QDRANT_URL,
                    upload_status="success"
                )
                db.add(history)
                db.commit()

                logger.info(f"Excel document updated in SQLite: id={excel_doc.id}, filename={request.file_name}, collection={request.collection_name}")

            except Exception as db_error:
                logger.error(f"Failed to update Excel document in SQLite: {db_error}")
                db.rollback()
                # Qdrant 업로드는 성공했으므로 에러를 던지지 않고 로그만 기록
        else:
            # 모든 배치가 실패한 경우 Document 삭제
            try:
                db.delete(excel_doc)
                db.commit()
                logger.warning(f"Deleted Excel document due to complete upload failure: id={excel_doc.id}")
            except Exception as del_error:
                logger.error(f"Failed to delete Excel document: {del_error}")
                db.rollback()

        return DynamicEmbeddingResponse(
            total=len(rows),
            success_count=success_count,
            failure_count=failure_count,
            results=results
        )

    except HTTPException:
        # Document 정리 (생성된 경우)
        if excel_doc and excel_doc.id:
            try:
                db.delete(excel_doc)
                db.commit()
            except Exception:
                db.rollback()
        raise
    except Exception as e:
        # Document 정리 (생성된 경우)
        if excel_doc and excel_doc.id:
            try:
                db.delete(excel_doc)
                db.commit()
            except Exception:
                db.rollback()
        print(f"[ERROR] Dynamic embedding failed: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Excel 임베딩 실패: {str(e)}"
        )


@router.post("/migrate/excel-collection/{collection_name}")
async def migrate_excel_collection(
    collection_name: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    기존 엑셀 컬렉션을 SQLite에 역으로 기록

    Qdrant에서 메타데이터를 추출하여 Document 및 QdrantUploadHistory 생성
    프롬프트 자동생성 모달에서 엑셀 문서가 표시되도록 함
    """
    try:
        # 1. 컬렉션 존재 확인
        collection_exists = await qdrant_service.collection_exists(collection_name)
        if not collection_exists:
            raise HTTPException(
                status_code=404,
                detail=f"Collection '{collection_name}'이 존재하지 않습니다"
            )

        # 2. 기존 마이그레이션 확인 제거 (Excel 문서가 있으면 마이그레이션 진행)
        # 기존 일반 문서 history가 있어도 Excel 문서는 별도로 마이그레이션 가능

        # 3. Qdrant에서 포인트 조회 (scroll)
        all_points = []
        offset = None
        while True:
            points, next_offset = await qdrant_service.client.scroll(
                collection_name=collection_name,
                limit=100,
                offset=offset,
                with_payload=True,
                with_vectors=False
            )
            all_points.extend(points)
            if next_offset is None:
                break
            offset = next_offset

        if not all_points:
            raise HTTPException(
                status_code=400,
                detail=f"Collection '{collection_name}'에 데이터가 없습니다"
            )

        # 4. Excel 포인트만 필터링 (row_index 있고 document_id 없는 포인트)
        excel_points = [
            p for p in all_points
            if (p.payload or {}).get("row_index") is not None
            and "document_id" not in (p.payload or {})
        ]

        if not excel_points:
            return {
                "success": True,
                "collection_name": collection_name,
                "migrated_documents": 0,
                "documents": [],
                "total_points": len(all_points),
                "message": "마이그레이션할 Excel 데이터가 없습니다 (이미 document_id가 있거나 Excel 데이터가 아님)"
            }

        # 5. filename 기준 그룹핑 (source_file 하위호환)
        grouped = {}
        for point in excel_points:
            payload = point.payload or {}
            # filename 우선, 없으면 source_file (하위호환)
            file_name = payload.get("filename") or payload.get("source_file", "unknown.xlsx")
            if file_name not in grouped:
                grouped[file_name] = {
                    "texts": [],
                    "vector_ids": [],
                    "row_indices": [],
                    "sample_payload": payload
                }
            grouped[file_name]["texts"].append(payload.get("text", ""))
            grouped[file_name]["vector_ids"].append(str(point.id))
            grouped[file_name]["row_indices"].append(payload.get("row_index", 0))

        # 6. 각 filename에 대해 Document 생성
        created_docs = []
        for file_name, data in grouped.items():
            from backend.services.excel_document_service import generate_preview

            # 메타정보 생성
            sample = data["sample_payload"]
            text_columns = [k for k in sample.keys() if k not in [
                "filename", "source_file", "row_index", "id", "tags", "headings", "text"
            ]]

            md_content = f"""# {file_name}

## 문서 정보
- **유형**: Excel 데이터 (마이그레이션됨)
- **총 행 수**: {len(data["texts"])}
- **컬렉션**: {collection_name}

## 메타데이터 컬럼
- {', '.join(text_columns) if text_columns else '없음'}

## 샘플링 안내
이 문서의 내용은 Qdrant 벡터 DB에 행 단위로 임베딩되어 있습니다.
프롬프트 자동생성 시 청크 기반 샘플링을 통해 의미론적으로 관련된 행들이 자동 추출됩니다.
"""

            doc = Document(
                task_id=f"excel-migrated-{uuid.uuid4().hex[:8]}",
                original_filename=file_name,
                file_type="xlsx",
                status="success",
                content_length=sum(len(t) for t in data["texts"]),
                content_preview=generate_preview(data["texts"][:5]),
                md_content=md_content,
                category=collection_name,
                parse_options={
                    "source_type": "excel",
                    "migrated": True,
                    "original_collection": collection_name
                }
            )
            db.add(doc)
            db.flush()

            # Qdrant payload에 document_id 추가
            point_ids = [
                int(vid) if vid.isdigit() else vid
                for vid in data["vector_ids"]
            ]
            await qdrant_service.client.set_payload(
                collection_name=collection_name,
                payload={"document_id": doc.id},
                points=point_ids
            )
            logger.info(f"Updated {len(point_ids)} Qdrant points with document_id={doc.id}")

            history = QdrantUploadHistory(
                document_id=doc.id,
                collection_name=collection_name,
                chunk_count=len(data["texts"]),
                vector_ids_json=json.dumps(data["vector_ids"]),
                qdrant_url=settings.QDRANT_URL,
                upload_status="success"
            )
            db.add(history)
            created_docs.append({
                "document_id": doc.id,
                "filename": file_name,
                "rows": len(data["texts"]),
                "qdrant_updated": True
            })

        db.commit()

        logger.info(f"Excel collection migrated: {collection_name}, {len(created_docs)} documents, Qdrant payloads updated")

        return {
            "success": True,
            "collection_name": collection_name,
            "migrated_documents": len(created_docs),
            "documents": created_docs,
            "total_points": len(all_points)
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to migrate Excel collection: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"마이그레이션 실패: {str(e)}"
        )
