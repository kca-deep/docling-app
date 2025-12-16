"""
Qdrant Vector DB 연동 API 라우트
컬렉션 가시성(visibility) 기반 접근 제어:
- public: 모든 사용자 접근 가능
- private: 소유자만 접근 가능
- shared: 소유자 + 허용된 사용자 접근 가능
"""
import uuid
import io
import json
import logging
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional, AsyncGenerator

from backend.services.qdrant_service import QdrantService
from backend.services.chunking_service import ChunkingService
from backend.services.embedding_service import EmbeddingService
from backend.services import document_crud, qdrant_history_crud, collection_crud
from backend.database import get_db
from backend.config.settings import settings
from backend.dependencies.auth import get_current_active_user, get_current_user_optional
from backend.models.user import User
from backend.models.document import Document
from backend.models.qdrant_upload_history import QdrantUploadHistory
from backend.models.schemas import (
    QdrantCollectionsResponse,
    QdrantCollectionCreateRequest,
    QdrantCollectionResponse,
    QdrantCollectionInfo,
    QdrantCollectionSettingsRequest,
    QdrantUploadRequest,
    QdrantUploadResponse,
    QdrantUploadResult,
    QdrantUploadProgressEvent,
    QdrantConfigResponse,
    DuplicateCheckRequest,
    DuplicateCheckResponse,
    DuplicateInfo,
    QAPreviewResponse,
    QAPreviewRow,
    QAEmbeddingRequest,
    QAEmbeddingResponse,
    QAEmbeddingResult,
    ExcelPreviewResponse,
    ExcelPreviewRow,
    ColumnMapping,
    DynamicEmbeddingRequest,
    DynamicEmbeddingResponse,
    DynamicEmbeddingResult,
    CollectionDocumentInfo,
    CollectionDocumentsResponse,
    DeleteDocumentRequest,
    DeleteDocumentResponse
)

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/qdrant",
    tags=["qdrant"]
    # 인증은 엔드포인트별로 처리
)

# 서비스 인스턴스 생성
qdrant_service = QdrantService(
    url=settings.QDRANT_URL,
    api_key=settings.QDRANT_API_KEY
)
chunking_service = ChunkingService(base_url=settings.DOCLING_CHUNKING_URL)
embedding_service = EmbeddingService(
    base_url=settings.EMBEDDING_URL,
    model=settings.EMBEDDING_MODEL
)


@router.get("/config", response_model=QdrantConfigResponse)
async def get_qdrant_config(
    current_user: User = Depends(get_current_active_user)
):
    """
    Qdrant 청킹 설정값 조회 API

    Returns:
        QdrantConfigResponse: 기본 chunk_size 및 chunk_overlap 설정값
    """
    return QdrantConfigResponse(
        default_chunk_size=settings.DEFAULT_CHUNK_SIZE,
        default_chunk_overlap=settings.DEFAULT_CHUNK_OVERLAP
    )


@router.get("/collections", response_model=QdrantCollectionsResponse)
async def get_collections(
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """
    Qdrant Collection 목록 조회 API

    사용자 권한에 따라 접근 가능한 컬렉션만 반환:
    - 비로그인: public 컬렉션만
    - 로그인: public + 소유 + 공유된(allowed) 컬렉션

    Returns:
        QdrantCollectionsResponse: Collection 목록

    Raises:
        HTTPException: Qdrant 서버 연결 실패 시
    """
    try:
        # 1. Qdrant에서 모든 컬렉션 조회
        qdrant_collections = await qdrant_service.get_collections()
        qdrant_names = [col.name for col in qdrant_collections]

        # 2. SQLite에서 접근 가능한 컬렉션 메타데이터 조회
        user_id = current_user.id if current_user else None
        accessible_metadata = collection_crud.get_accessible_collections(
            db=db,
            user_id=user_id,
            qdrant_collection_names=qdrant_names
        )

        # 3. 메타데이터를 딕셔너리로 변환 (빠른 조회용)
        metadata_map = {col.collection_name: col for col in accessible_metadata}

        # 4. Qdrant 데이터와 SQLite 메타데이터 병합
        result_collections = []
        for qdrant_col in qdrant_collections:
            if qdrant_col.name in metadata_map:
                meta = metadata_map[qdrant_col.name]
                result_collections.append(QdrantCollectionInfo(
                    name=qdrant_col.name,
                    documents_count=qdrant_col.documents_count,
                    points_count=qdrant_col.points_count,
                    vector_size=qdrant_col.vector_size,
                    distance=qdrant_col.distance,
                    visibility=meta.visibility,
                    description=meta.description,
                    owner_id=meta.owner_id,
                    is_owner=user_id is not None and meta.owner_id == user_id
                ))

        return QdrantCollectionsResponse(
            collections=result_collections
        )

    except Exception as e:
        logger.error(f"Failed to get collections: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Qdrant collection 목록 조회 실패: {str(e)}"
        )


@router.post("/collections", response_model=QdrantCollectionResponse)
async def create_collection(
    request: QdrantCollectionCreateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Qdrant Collection 생성 API

    Args:
        request: Collection 생성 요청 (visibility, description 포함)
        db: DB 세션
        current_user: 현재 로그인 사용자 (소유자로 지정)

    Returns:
        QdrantCollectionResponse: 생성 결과

    Raises:
        HTTPException: Collection 생성 실패 시
    """
    try:
        # Collection 이름 검증
        if not request.collection_name or not request.collection_name.strip():
            raise HTTPException(
                status_code=400,
                detail="Collection 이름은 필수입니다"
            )

        # Vector size 검증
        if request.vector_size <= 0:
            raise HTTPException(
                status_code=400,
                detail="Vector size는 양수여야 합니다"
            )

        # Distance metric 검증
        valid_distances = ["Cosine", "Euclidean", "Dot"]
        if request.distance not in valid_distances:
            raise HTTPException(
                status_code=400,
                detail=f"Distance metric은 {valid_distances} 중 하나여야 합니다"
            )

        # Visibility 검증
        valid_visibilities = ["public", "private", "shared"]
        if request.visibility not in valid_visibilities:
            raise HTTPException(
                status_code=400,
                detail=f"Visibility는 {valid_visibilities} 중 하나여야 합니다"
            )

        # 1. Qdrant에 Collection 생성
        await qdrant_service.create_collection(
            collection_name=request.collection_name,
            vector_size=request.vector_size,
            distance=request.distance
        )

        # 2. SQLite에 메타데이터 저장
        try:
            collection_crud.create_collection(
                db=db,
                collection_name=request.collection_name,
                owner_id=current_user.id,
                visibility=request.visibility,
                description=request.description
            )
            logger.info(f"Created collection metadata: {request.collection_name} (owner={current_user.id})")
        except Exception as e:
            # SQLite 저장 실패 시에도 Qdrant에는 생성되었으므로 경고만 로깅
            logger.warning(f"Failed to save collection metadata to SQLite: {e}")

        return QdrantCollectionResponse(
            success=True,
            collection_name=request.collection_name,
            message=f"Collection '{request.collection_name}'이 성공적으로 생성되었습니다"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create collection: {e}")

        # 이미 존재하는 경우
        if "이미 존재합니다" in str(e):
            raise HTTPException(
                status_code=409,
                detail=str(e)
            )

        raise HTTPException(
            status_code=500,
            detail=f"Collection 생성 실패: {str(e)}"
        )


@router.delete("/collections/{collection_name}", response_model=QdrantCollectionResponse)
async def delete_collection(
    collection_name: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Qdrant Collection 삭제 API

    소유자 또는 관리자만 삭제 가능

    Args:
        collection_name: 삭제할 Collection 이름
        db: DB 세션
        current_user: 현재 로그인 사용자

    Returns:
        QdrantCollectionResponse: 삭제 결과

    Raises:
        HTTPException: Collection 삭제 실패 시
    """
    try:
        # Collection 이름 검증
        if not collection_name or not collection_name.strip():
            raise HTTPException(
                status_code=400,
                detail="Collection 이름은 필수입니다"
            )

        # 소유권 확인 (관리자는 모든 컬렉션 삭제 가능)
        collection_meta = collection_crud.get_by_name(db, collection_name)
        if collection_meta:
            is_owner = collection_meta.owner_id == current_user.id
            is_admin = current_user.role == "admin"
            if not is_owner and not is_admin:
                raise HTTPException(
                    status_code=403,
                    detail="컬렉션 삭제 권한이 없습니다. 소유자만 삭제할 수 있습니다."
                )

        # 1. Qdrant에서 Collection 삭제
        await qdrant_service.delete_collection(collection_name)

        # 2. SQLite에서 메타데이터 삭제
        try:
            collection_crud.delete_collection(db, collection_name)
            logger.info(f"Deleted collection metadata: {collection_name}")
        except Exception as e:
            # SQLite 삭제 실패 시에도 Qdrant에서는 삭제되었으므로 경고만 로깅
            logger.warning(f"Failed to delete collection metadata from SQLite: {e}")

        return QdrantCollectionResponse(
            success=True,
            collection_name=collection_name,
            message=f"Collection '{collection_name}'이 성공적으로 삭제되었습니다"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete collection: {e}")

        # 존재하지 않는 경우
        if "존재하지 않습니다" in str(e):
            raise HTTPException(
                status_code=404,
                detail=str(e)
            )

        raise HTTPException(
            status_code=500,
            detail=f"Collection 삭제 실패: {str(e)}"
        )


@router.patch("/collections/{collection_name}/settings", response_model=QdrantCollectionResponse)
async def update_collection_settings(
    collection_name: str,
    request: QdrantCollectionSettingsRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    컬렉션 설정 변경 API

    소유자만 설정 변경 가능

    Args:
        collection_name: 컬렉션 이름
        request: 설정 변경 요청 (visibility, description, allowed_users)
        db: DB 세션
        current_user: 현재 로그인 사용자

    Returns:
        QdrantCollectionResponse: 변경 결과

    Raises:
        HTTPException: 설정 변경 실패 시
    """
    try:
        # 컬렉션 메타데이터 조회
        collection_meta = collection_crud.get_by_name(db, collection_name)
        if not collection_meta:
            raise HTTPException(
                status_code=404,
                detail=f"컬렉션 '{collection_name}'의 메타데이터를 찾을 수 없습니다"
            )

        # 소유권 확인
        if collection_meta.owner_id != current_user.id:
            raise HTTPException(
                status_code=403,
                detail="컬렉션 설정 변경 권한이 없습니다. 소유자만 변경할 수 있습니다."
            )

        # Visibility 검증
        if request.visibility:
            valid_visibilities = ["public", "private", "shared"]
            if request.visibility not in valid_visibilities:
                raise HTTPException(
                    status_code=400,
                    detail=f"Visibility는 {valid_visibilities} 중 하나여야 합니다"
                )

        # 설정 업데이트
        updated = collection_crud.update_settings(
            db=db,
            collection_name=collection_name,
            description=request.description,
            visibility=request.visibility,
            allowed_users=request.allowed_users
        )

        if not updated:
            raise HTTPException(
                status_code=500,
                detail="설정 변경에 실패했습니다"
            )

        logger.info(f"Updated collection settings: {collection_name}")

        return QdrantCollectionResponse(
            success=True,
            collection_name=collection_name,
            message=f"Collection '{collection_name}' 설정이 변경되었습니다"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update collection settings: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"설정 변경 실패: {str(e)}"
        )


@router.get("/collections/{collection_name}/documents", response_model=CollectionDocumentsResponse)
async def get_collection_documents(
    collection_name: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    컬렉션 내 문서 목록 조회 API

    소유자 또는 관리자만 조회 가능

    Args:
        collection_name: Collection 이름
        db: DB 세션
        current_user: 현재 로그인 사용자

    Returns:
        CollectionDocumentsResponse: 문서 목록
    """
    try:
        # 컬렉션 존재 확인
        exists = await qdrant_service.collection_exists(collection_name)
        if not exists:
            raise HTTPException(
                status_code=404,
                detail=f"Collection '{collection_name}'이 존재하지 않습니다"
            )

        # 권한 확인 (소유자 또는 관리자)
        collection_meta = collection_crud.get_by_name(db, collection_name)
        if collection_meta:
            is_owner = collection_meta.owner_id == current_user.id
            is_admin = current_user.role == "admin"
            if not is_owner and not is_admin:
                raise HTTPException(
                    status_code=403,
                    detail="문서 목록 조회 권한이 없습니다"
                )

        # Qdrant에서 문서 목록 조회
        documents = await qdrant_service.get_documents_in_collection(collection_name)

        return CollectionDocumentsResponse(
            collection_name=collection_name,
            total_documents=len(documents),
            documents=[CollectionDocumentInfo(**doc) for doc in documents]
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get collection documents: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"문서 목록 조회 실패: {str(e)}"
        )


@router.delete("/collections/{collection_name}/documents", response_model=DeleteDocumentResponse)
async def delete_collection_documents(
    collection_name: str,
    request: DeleteDocumentRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    컬렉션에서 문서 삭제 API

    document_ids 또는 source_files 중 하나 이상 필수
    소유자 또는 관리자만 삭제 가능

    Args:
        collection_name: Collection 이름
        request: 삭제 요청 (document_ids 또는 source_files)
        db: DB 세션
        current_user: 현재 로그인 사용자

    Returns:
        DeleteDocumentResponse: 삭제 결과
    """
    try:
        # 요청 유효성 검사
        if not request.document_ids and not request.source_files:
            raise HTTPException(
                status_code=400,
                detail="document_ids 또는 source_files 중 하나 이상 필요합니다"
            )

        # 컬렉션 존재 확인
        exists = await qdrant_service.collection_exists(collection_name)
        if not exists:
            raise HTTPException(
                status_code=404,
                detail=f"Collection '{collection_name}'이 존재하지 않습니다"
            )

        # 권한 확인 (소유자 또는 관리자)
        collection_meta = collection_crud.get_by_name(db, collection_name)
        if collection_meta:
            is_owner = collection_meta.owner_id == current_user.id
            is_admin = current_user.role == "admin"
            if not is_owner and not is_admin:
                raise HTTPException(
                    status_code=403,
                    detail="문서 삭제 권한이 없습니다"
                )

        total_deleted = 0

        # document_id로 삭제 (일반 문서)
        if request.document_ids:
            for doc_id in request.document_ids:
                deleted = await qdrant_service.delete_document_points(collection_name, doc_id)
                total_deleted += deleted

                # SQLite 업로드 이력도 삭제
                qdrant_history_crud.delete_by_document_and_collection(
                    db, doc_id, collection_name
                )

        # source_file로 삭제 (Excel)
        if request.source_files:
            for source_file in request.source_files:
                deleted = await qdrant_service.delete_excel_points(collection_name, source_file)
                total_deleted += deleted

        logger.info(f"Deleted {total_deleted} points from collection '{collection_name}'")

        return DeleteDocumentResponse(
            success=True,
            deleted_count=total_deleted,
            message=f"{total_deleted}개의 벡터가 삭제되었습니다"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete collection documents: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"문서 삭제 실패: {str(e)}"
        )


@router.post("/check-duplicates", response_model=DuplicateCheckResponse)
async def check_duplicates(
    request: DuplicateCheckRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    문서 중복 업로드 확인 API

    지정한 문서들이 이미 해당 Collection에 업로드되었는지 확인합니다.

    Args:
        request: 중복 확인 요청
        db: DB 세션

    Returns:
        DuplicateCheckResponse: 중복 문서 목록
    """
    duplicates = []
    new_documents = []

    for document_id in request.document_ids:
        # 이미 업로드된 이력 확인
        existing = qdrant_history_crud.check_document_uploaded_to_collection(
            db=db,
            document_id=document_id,
            collection_name=request.collection_name
        )

        if existing:
            # 중복 문서
            document = document_crud.get_document_by_id(db, document_id)
            duplicates.append(DuplicateInfo(
                document_id=document_id,
                filename=document.original_filename if document else "Unknown",
                uploaded_at=existing.uploaded_at.isoformat()
            ))
        else:
            # 신규 문서
            new_documents.append(document_id)

    return DuplicateCheckResponse(
        has_duplicates=len(duplicates) > 0,
        duplicates=duplicates,
        new_documents=new_documents
    )


@router.post("/upload", response_model=QdrantUploadResponse)
async def upload_documents(
    request: QdrantUploadRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    문서를 Qdrant에 임베딩 및 업로드

    Args:
        request: 업로드 요청
        db: DB 세션

    Returns:
        QdrantUploadResponse: 업로드 결과

    Raises:
        HTTPException: 업로드 실패 시
    """
    results = []
    success_count = 0
    failure_count = 0

    try:
        # Collection 존재 확인
        collection_exists = await qdrant_service.collection_exists(request.collection_name)
        if not collection_exists:
            raise HTTPException(
                status_code=404,
                detail=f"Collection '{request.collection_name}'이 존재하지 않습니다"
            )

        # 각 문서 처리
        for document_id in request.document_ids:
            try:
                # 1. DB에서 문서 조회
                document = document_crud.get_document_by_id(db, document_id)
                if not document or not document.md_content:
                    results.append(QdrantUploadResult(
                        document_id=document_id,
                        filename=document.original_filename if document else "Unknown",
                        success=False,
                        error="문서를 찾을 수 없거나 markdown 내용이 없습니다"
                    ))
                    failure_count += 1
                    continue

                print(f"[INFO] Processing document {document_id}: {document.original_filename}")

                # 2. Markdown 청킹
                chunks = await chunking_service.chunk_markdown(
                    markdown_content=document.md_content,
                    max_tokens=request.chunk_size,
                    filename=document.original_filename
                )

                if not chunks:
                    results.append(QdrantUploadResult(
                        document_id=document_id,
                        filename=document.original_filename,
                        success=False,
                        error="청킹 결과가 없습니다"
                    ))
                    failure_count += 1
                    continue

                # 3. 청크 텍스트 추출
                chunk_texts = [chunk.get('text', '') for chunk in chunks]

                # 4. 임베딩 생성
                embeddings = await embedding_service.get_embeddings(chunk_texts)

                if len(embeddings) != len(chunk_texts):
                    results.append(QdrantUploadResult(
                        document_id=document_id,
                        filename=document.original_filename,
                        success=False,
                        error="임베딩 개수와 청크 개수가 일치하지 않습니다"
                    ))
                    failure_count += 1
                    continue

                # 5. 메타데이터 생성
                metadata_list = []
                for i, chunk in enumerate(chunks):
                    metadata_list.append({
                        "document_id": document_id,
                        "filename": document.original_filename,
                        "chunk_index": i,
                        "num_tokens": chunk.get('num_tokens', 0),
                        "headings": chunk.get('headings') or []  # None 안전 처리
                    })

                # 6. Qdrant에 벡터 업로드
                vector_ids = await qdrant_service.upsert_vectors(
                    collection_name=request.collection_name,
                    vectors=embeddings,
                    texts=chunk_texts,
                    metadata_list=metadata_list
                )

                # 7. 업로드 이력 저장
                qdrant_history_crud.create_upload_history(
                    db=db,
                    document_id=document_id,
                    collection_name=request.collection_name,
                    chunk_count=len(chunks),
                    vector_ids=vector_ids,
                    qdrant_url=settings.QDRANT_URL,
                    upload_status="success"
                )

                # 성공 결과 추가
                results.append(QdrantUploadResult(
                    document_id=document_id,
                    filename=document.original_filename,
                    success=True,
                    chunk_count=len(chunks),
                    vector_ids=vector_ids
                ))
                success_count += 1

                print(f"[INFO] Successfully uploaded document {document_id} with {len(chunks)} chunks")

            except Exception as e:
                print(f"[ERROR] Failed to upload document {document_id}: {e}")

                # 실패 이력 저장
                try:
                    document = document_crud.get_document_by_id(db, document_id)
                    qdrant_history_crud.create_upload_history(
                        db=db,
                        document_id=document_id,
                        collection_name=request.collection_name,
                        chunk_count=0,
                        vector_ids=[],
                        qdrant_url=settings.QDRANT_URL,
                        upload_status="failure",
                        error_message=str(e)
                    )
                except:
                    pass

                results.append(QdrantUploadResult(
                    document_id=document_id,
                    filename=document.original_filename if document else "Unknown",
                    success=False,
                    error=str(e)
                ))
                failure_count += 1

        return QdrantUploadResponse(
            total=len(request.document_ids),
            success_count=success_count,
            failure_count=failure_count,
            results=results
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Upload process failed: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"업로드 프로세스 실패: {str(e)}"
        )


@router.post("/upload/stream")
async def upload_documents_stream(
    request: QdrantUploadRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    문서를 Qdrant에 임베딩 및 업로드 (SSE 스트리밍)

    실시간 진행률을 SSE(Server-Sent Events)로 전송합니다.

    Args:
        request: 업로드 요청
        db: DB 세션

    Returns:
        StreamingResponse: SSE 이벤트 스트림
    """

    async def generate_events() -> AsyncGenerator[str, None]:
        """SSE 이벤트 생성기"""
        results = []
        success_count = 0
        failure_count = 0
        total_docs = len(request.document_ids)

        def send_event(event: QdrantUploadProgressEvent) -> str:
            """SSE 포맷으로 이벤트 전송"""
            return f"data: {event.model_dump_json()}\n\n"

        try:
            # Collection 존재 확인
            collection_exists = await qdrant_service.collection_exists(request.collection_name)
            if not collection_exists:
                yield send_event(QdrantUploadProgressEvent(
                    event_type="error",
                    error=f"Collection '{request.collection_name}'이 존재하지 않습니다",
                    total_docs=total_docs
                ))
                return

            # 각 문서 처리
            for doc_index, document_id in enumerate(request.document_ids):
                try:
                    # 1. DB에서 문서 조회
                    document = document_crud.get_document_by_id(db, document_id)
                    if not document or not document.md_content:
                        yield send_event(QdrantUploadProgressEvent(
                            event_type="error",
                            document_id=document_id,
                            filename=document.original_filename if document else "Unknown",
                            error="문서를 찾을 수 없거나 markdown 내용이 없습니다",
                            current_doc_index=doc_index + 1,
                            total_docs=total_docs,
                            success_count=success_count,
                            failure_count=failure_count + 1
                        ))
                        failure_count += 1
                        results.append(QdrantUploadResult(
                            document_id=document_id,
                            filename=document.original_filename if document else "Unknown",
                            success=False,
                            error="문서를 찾을 수 없거나 markdown 내용이 없습니다"
                        ))
                        continue

                    filename = document.original_filename

                    # 2. 청킹 시작
                    yield send_event(QdrantUploadProgressEvent(
                        event_type="progress",
                        document_id=document_id,
                        filename=filename,
                        phase="chunking",
                        progress=10,
                        current_doc_index=doc_index + 1,
                        total_docs=total_docs,
                        success_count=success_count,
                        failure_count=failure_count
                    ))

                    chunks = await chunking_service.chunk_markdown(
                        markdown_content=document.md_content,
                        max_tokens=request.chunk_size,
                        filename=filename
                    )

                    if not chunks:
                        yield send_event(QdrantUploadProgressEvent(
                            event_type="error",
                            document_id=document_id,
                            filename=filename,
                            error="청킹 결과가 없습니다",
                            current_doc_index=doc_index + 1,
                            total_docs=total_docs,
                            success_count=success_count,
                            failure_count=failure_count + 1
                        ))
                        failure_count += 1
                        results.append(QdrantUploadResult(
                            document_id=document_id,
                            filename=filename,
                            success=False,
                            error="청킹 결과가 없습니다"
                        ))
                        continue

                    chunk_texts = [chunk.get('text', '') for chunk in chunks]

                    # 3. 임베딩 시작
                    yield send_event(QdrantUploadProgressEvent(
                        event_type="progress",
                        document_id=document_id,
                        filename=filename,
                        phase="embedding",
                        progress=40,
                        current_doc_index=doc_index + 1,
                        total_docs=total_docs,
                        chunk_count=len(chunks),
                        success_count=success_count,
                        failure_count=failure_count
                    ))

                    embeddings = await embedding_service.get_embeddings(chunk_texts)

                    if len(embeddings) != len(chunk_texts):
                        yield send_event(QdrantUploadProgressEvent(
                            event_type="error",
                            document_id=document_id,
                            filename=filename,
                            error="임베딩 개수와 청크 개수가 일치하지 않습니다",
                            current_doc_index=doc_index + 1,
                            total_docs=total_docs,
                            success_count=success_count,
                            failure_count=failure_count + 1
                        ))
                        failure_count += 1
                        results.append(QdrantUploadResult(
                            document_id=document_id,
                            filename=filename,
                            success=False,
                            error="임베딩 개수와 청크 개수가 일치하지 않습니다"
                        ))
                        continue

                    # 4. Qdrant 업로드 시작
                    yield send_event(QdrantUploadProgressEvent(
                        event_type="progress",
                        document_id=document_id,
                        filename=filename,
                        phase="uploading",
                        progress=70,
                        current_doc_index=doc_index + 1,
                        total_docs=total_docs,
                        chunk_count=len(chunks),
                        success_count=success_count,
                        failure_count=failure_count
                    ))

                    # 메타데이터 생성
                    metadata_list = []
                    for i, chunk in enumerate(chunks):
                        metadata_list.append({
                            "document_id": document_id,
                            "filename": filename,
                            "chunk_index": i,
                            "num_tokens": chunk.get('num_tokens', 0),
                            "headings": chunk.get('headings') or []
                        })

                    # Qdrant에 벡터 업로드
                    vector_ids = await qdrant_service.upsert_vectors(
                        collection_name=request.collection_name,
                        vectors=embeddings,
                        texts=chunk_texts,
                        metadata_list=metadata_list
                    )

                    # 업로드 이력 저장
                    qdrant_history_crud.create_upload_history(
                        db=db,
                        document_id=document_id,
                        collection_name=request.collection_name,
                        chunk_count=len(chunks),
                        vector_ids=vector_ids,
                        qdrant_url=settings.QDRANT_URL,
                        upload_status="success"
                    )

                    success_count += 1
                    results.append(QdrantUploadResult(
                        document_id=document_id,
                        filename=filename,
                        success=True,
                        chunk_count=len(chunks),
                        vector_ids=vector_ids
                    ))

                    # 문서 완료 이벤트
                    yield send_event(QdrantUploadProgressEvent(
                        event_type="document_complete",
                        document_id=document_id,
                        filename=filename,
                        phase="completed",
                        progress=100,
                        current_doc_index=doc_index + 1,
                        total_docs=total_docs,
                        chunk_count=len(chunks),
                        vector_ids=vector_ids,
                        success_count=success_count,
                        failure_count=failure_count
                    ))

                except Exception as e:
                    logger.error(f"Failed to upload document {document_id}: {e}")

                    # 실패 이력 저장
                    try:
                        document = document_crud.get_document_by_id(db, document_id)
                        qdrant_history_crud.create_upload_history(
                            db=db,
                            document_id=document_id,
                            collection_name=request.collection_name,
                            chunk_count=0,
                            vector_ids=[],
                            qdrant_url=settings.QDRANT_URL,
                            upload_status="failure",
                            error_message=str(e)
                        )
                    except:
                        pass

                    failure_count += 1
                    results.append(QdrantUploadResult(
                        document_id=document_id,
                        filename=document.original_filename if document else "Unknown",
                        success=False,
                        error=str(e)
                    ))

                    yield send_event(QdrantUploadProgressEvent(
                        event_type="error",
                        document_id=document_id,
                        filename=document.original_filename if document else "Unknown",
                        error=str(e),
                        current_doc_index=doc_index + 1,
                        total_docs=total_docs,
                        success_count=success_count,
                        failure_count=failure_count
                    ))

            # 완료 이벤트
            yield send_event(QdrantUploadProgressEvent(
                event_type="done",
                progress=100,
                current_doc_index=total_docs,
                total_docs=total_docs,
                success_count=success_count,
                failure_count=failure_count
            ))

        except Exception as e:
            logger.error(f"Upload stream failed: {e}")
            yield send_event(QdrantUploadProgressEvent(
                event_type="error",
                error=f"업로드 프로세스 실패: {str(e)}",
                total_docs=total_docs,
                success_count=success_count,
                failure_count=failure_count
            ))

    return StreamingResponse(
        generate_events(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


# ==================== Q&A Excel Embedding Endpoints ====================

@router.post("/qa/preview", response_model=QAPreviewResponse)
async def preview_qa_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user)
):
    """
    Q&A Excel 파일 미리보기 API

    Args:
        file: 업로드된 Excel 파일 (.xlsx)

    Returns:
        QAPreviewResponse: 미리보기 데이터
    """
    try:
        import openpyxl

        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="Excel 파일(.xlsx, .xls)만 지원합니다"
            )

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        headers = [cell.value for cell in ws[1] if cell.value]
        print(f"[INFO] Excel headers: {headers}")

        required_headers = ['question', 'answer_text']
        header_lower = [h.lower() if h else '' for h in headers]

        for req in required_headers:
            if req not in header_lower:
                raise HTTPException(
                    status_code=400,
                    detail=f"필수 컬럼 '{req}'이(가) 없습니다. 현재 컬럼: {headers}"
                )

        header_map = {h.lower(): i for i, h in enumerate(headers) if h}

        rows = []
        for row_num in range(2, ws.max_row + 1):
            row_values = [ws.cell(row=row_num, column=i+1).value for i in range(len(headers))]

            question = row_values[header_map.get('question', 1)] or ''
            answer_text = row_values[header_map.get('answer_text', 2)] or ''

            if not question.strip() and not answer_text.strip():
                continue

            faq_id = row_values[header_map.get('faq_id', 0)] if 'faq_id' in header_map else f"FAQ-{row_num-1:04d}"
            tags_raw = row_values[header_map.get('tag', -1)] if 'tag' in header_map else ''
            tags = [t.strip() for t in (tags_raw or '').split(',')] if tags_raw else []
            policy_anchor = row_values[header_map.get('policy_anchor', -1)] if 'policy_anchor' in header_map else None
            source = row_values[header_map.get('source', -1)] if 'source' in header_map else None

            rows.append(QAPreviewRow(
                row_index=row_num - 2,
                faq_id=str(faq_id) if faq_id else f"FAQ-{row_num-1:04d}",
                question=str(question),
                answer_text=str(answer_text),
                tags=tags,
                policy_anchor=str(policy_anchor) if policy_anchor else None,
                source=str(source) if source else None
            ))

        wb.close()

        print(f"[INFO] Parsed {len(rows)} Q&A rows from Excel")

        return QAPreviewResponse(
            total_rows=len(rows),
            headers=headers,
            preview_rows=rows,
            file_name=file.filename
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Failed to parse Excel file: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Excel 파일 파싱 실패: {str(e)}"
        )


@router.post("/qa/embed", response_model=QAEmbeddingResponse)
async def embed_qa_rows(
    request: QAEmbeddingRequest,
    current_user: User = Depends(get_current_active_user)
):
    """
    Q&A 행별 임베딩 및 Qdrant 업로드 API

    Args:
        request: Q&A 임베딩 요청

    Returns:
        QAEmbeddingResponse: 임베딩 결과
    """
    results = []
    success_count = 0
    failure_count = 0

    try:
        collection_exists = await qdrant_service.collection_exists(request.collection_name)
        if not collection_exists:
            raise HTTPException(
                status_code=404,
                detail=f"Collection '{request.collection_name}'이 존재하지 않습니다"
            )

        batch_size = settings.UPLOAD_BATCH_SIZE
        rows = request.rows

        for batch_start in range(0, len(rows), batch_size):
            batch_end = min(batch_start + batch_size, len(rows))
            batch_rows = rows[batch_start:batch_end]

            try:
                texts = [f"질문: {row.question}\n답변: {row.answer_text}" for row in batch_rows]

                embeddings = await embedding_service.get_embeddings(texts)

                metadata_list = []
                for row in batch_rows:
                    metadata_list.append({
                        "faq_id": row.faq_id,
                        "question": row.question,
                        "answer_text": row.answer_text,
                        "tags": row.tags,
                        "policy_anchor": row.policy_anchor or "",
                        "source": row.source or "",
                        "row_index": row.row_index
                    })

                vector_ids = await qdrant_service.upsert_vectors(
                    collection_name=request.collection_name,
                    vectors=embeddings,
                    texts=texts,
                    metadata_list=metadata_list
                )

                for i, row in enumerate(batch_rows):
                    results.append(QAEmbeddingResult(
                        row_index=row.row_index,
                        faq_id=row.faq_id,
                        success=True,
                        vector_id=vector_ids[i] if i < len(vector_ids) else None
                    ))
                    success_count += 1

                print(f"[INFO] Embedded batch {batch_start+1}-{batch_end}")

            except Exception as e:
                print(f"[ERROR] Failed to embed batch {batch_start+1}-{batch_end}: {e}")
                for row in batch_rows:
                    results.append(QAEmbeddingResult(
                        row_index=row.row_index,
                        faq_id=row.faq_id,
                        success=False,
                        error=str(e)
                    ))
                    failure_count += 1

        return QAEmbeddingResponse(
            total=len(rows),
            success_count=success_count,
            failure_count=failure_count,
            results=results
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Q&A embedding failed: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Q&A 임베딩 실패: {str(e)}"
        )


# ==================== Dynamic Excel Embedding Endpoints ====================

# 스마트 컬럼 감지 규칙
TEXT_CANDIDATES = ['full_text', 'content', 'text', 'answer_text', 'body', 'description']
ID_CANDIDATES = ['id', 'faq_id', 'law_id', 'doc_id', 'item_id', 'code']
TAG_CANDIDATES = ['tag', 'tags', 'category', 'categories', 'label', 'labels']
QA_QUESTION_CANDIDATES = ['question', 'q', 'query', 'title']
QA_ANSWER_CANDIDATES = ['answer', 'answer_text', 'a', 'response', 'content']
# headings 컬럼 후보 (문서명, 페이지 등 참조문서 표시에 사용)
HEADING_SOURCE_CANDIDATES = ['source', 'document', 'doc_name', 'file', 'filename', 'document_name', 'ref', 'reference']
HEADING_PAGE_CANDIDATES = ['page', 'page_number', 'page_no', 'pg', 'section', 'chapter']


def detect_column_mapping(headers: List[str]) -> dict:
    """헤더를 분석하여 컬럼 매핑 자동 감지"""
    headers_lower = [h.lower() if h else '' for h in headers]

    detected = {
        "id_column": None,
        "text_columns": [],
        "tag_column": None,
        "is_qa_format": False,
        "question_column": None,
        "answer_column": None,
        "heading_columns": []  # 참조문서 표시용 컬럼들
    }

    # ID 컬럼 감지
    for candidate in ID_CANDIDATES:
        for i, h in enumerate(headers_lower):
            if candidate in h:
                detected["id_column"] = headers[i]
                break
        if detected["id_column"]:
            break

    # Q&A 패턴 감지
    for q_candidate in QA_QUESTION_CANDIDATES:
        for i, h in enumerate(headers_lower):
            if q_candidate == h or h.endswith(q_candidate):
                detected["question_column"] = headers[i]
                break
        if detected["question_column"]:
            break

    for a_candidate in QA_ANSWER_CANDIDATES:
        for i, h in enumerate(headers_lower):
            if a_candidate == h or h.endswith(a_candidate):
                detected["answer_column"] = headers[i]
                break
        if detected["answer_column"]:
            break

    if detected["question_column"] and detected["answer_column"]:
        detected["is_qa_format"] = True
        detected["text_columns"] = [detected["question_column"], detected["answer_column"]]
    else:
        # 일반 텍스트 컬럼 감지
        for candidate in TEXT_CANDIDATES:
            for i, h in enumerate(headers_lower):
                if candidate in h:
                    detected["text_columns"].append(headers[i])

    # 태그 컬럼 감지
    for candidate in TAG_CANDIDATES:
        for i, h in enumerate(headers_lower):
            if candidate == h:
                detected["tag_column"] = headers[i]
                break
        if detected["tag_column"]:
            break

    # headings 컬럼 감지 (소스/문서명 + 페이지/섹션 순서로)
    heading_source = None
    heading_page = None

    for candidate in HEADING_SOURCE_CANDIDATES:
        for i, h in enumerate(headers_lower):
            if candidate == h or candidate in h:
                heading_source = headers[i]
                break
        if heading_source:
            break

    for candidate in HEADING_PAGE_CANDIDATES:
        for i, h in enumerate(headers_lower):
            if candidate == h or candidate in h:
                heading_page = headers[i]
                break
        if heading_page:
            break

    # 감지된 컬럼들을 heading_columns에 추가 (소스 먼저, 페이지 다음)
    if heading_source:
        detected["heading_columns"].append(heading_source)
    if heading_page:
        detected["heading_columns"].append(heading_page)

    return detected


@router.post("/excel/preview", response_model=ExcelPreviewResponse)
async def preview_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user)
):
    """
    Excel 파일 미리보기 및 스마트 컬럼 감지 API

    Args:
        file: 업로드된 Excel 파일 (.xlsx)

    Returns:
        ExcelPreviewResponse: 미리보기 데이터 및 감지된 매핑
    """
    try:
        import openpyxl

        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="Excel 파일(.xlsx, .xls)만 지원합니다"
            )

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        headers = [cell.value for cell in ws[1] if cell.value]
        print(f"[INFO] Excel headers: {headers}")

        # 스마트 컬럼 매핑 감지
        detected_mapping = detect_column_mapping(headers)
        print(f"[INFO] Detected mapping: {detected_mapping}")

        # 모든 행 읽기
        rows = []
        for row_num in range(2, ws.max_row + 1):
            row_data = {}
            has_content = False

            for col_num, header in enumerate(headers):
                cell_value = ws.cell(row=row_num, column=col_num + 1).value
                if cell_value is not None:
                    row_data[header] = str(cell_value) if cell_value else ""
                    if str(cell_value).strip():
                        has_content = True
                else:
                    row_data[header] = ""

            if has_content:
                rows.append(ExcelPreviewRow(
                    row_index=row_num - 2,
                    data=row_data
                ))

        wb.close()

        print(f"[INFO] Parsed {len(rows)} rows from Excel")

        return ExcelPreviewResponse(
            total_rows=len(rows),
            headers=headers,
            preview_rows=rows,
            file_name=file.filename,
            detected_mapping=detected_mapping
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Failed to parse Excel file: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Excel 파일 파싱 실패: {str(e)}"
        )


@router.post("/excel/embed", response_model=DynamicEmbeddingResponse)
async def embed_excel_dynamic(
    request: DynamicEmbeddingRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    동적 컬럼 매핑을 사용한 Excel 임베딩 API

    Args:
        request: 동적 임베딩 요청 (컬럼 매핑 포함)

    Returns:
        DynamicEmbeddingResponse: 임베딩 결과
    """
    results = []
    success_count = 0
    failure_count = 0
    vector_ids_all = []  # SQLite 기록용 전체 vector ID
    texts_all = []  # 미리보기용 텍스트

    try:
        # Collection 존재 확인
        collection_exists = await qdrant_service.collection_exists(request.collection_name)
        if not collection_exists:
            raise HTTPException(
                status_code=404,
                detail=f"Collection '{request.collection_name}'이 존재하지 않습니다"
            )

        mapping = request.mapping
        batch_size = settings.UPLOAD_BATCH_SIZE
        rows = request.rows

        for batch_start in range(0, len(rows), batch_size):
            batch_end = min(batch_start + batch_size, len(rows))
            batch_rows = rows[batch_start:batch_end]

            try:
                # 임베딩 텍스트 생성
                texts = []
                for row in batch_rows:
                    if mapping.text_template:
                        # 템플릿 사용
                        text = mapping.text_template
                        for key, value in row.data.items():
                            text = text.replace(f"{{{key}}}", str(value) if value else "")
                    else:
                        # 텍스트 컬럼 연결
                        text_parts = []
                        for col in mapping.text_columns:
                            if col in row.data and row.data[col]:
                                text_parts.append(str(row.data[col]))
                        text = "\n".join(text_parts)

                    texts.append(text)

                # 임베딩 생성
                embeddings = await embedding_service.get_embeddings(texts)

                # 메타데이터 생성
                metadata_list = []
                for row in batch_rows:
                    metadata = {
                        "source_file": request.file_name,
                        "row_index": row.row_index
                    }

                    # ID 컬럼
                    if mapping.id_column and mapping.id_column in row.data:
                        metadata["id"] = row.data[mapping.id_column]

                    # 태그 컬럼
                    if mapping.tag_column and mapping.tag_column in row.data:
                        tag_value = row.data[mapping.tag_column]
                        if tag_value:
                            metadata["tags"] = [t.strip() for t in str(tag_value).split(',')]

                    # 메타데이터 컬럼들
                    for col in mapping.metadata_columns:
                        if col in row.data:
                            metadata[col] = row.data[col]

                    # 텍스트 컬럼들도 메타데이터에 저장
                    for col in mapping.text_columns:
                        if col in row.data:
                            metadata[col] = row.data[col]

                    # headings 생성 (참조문서 표시용)
                    if mapping.heading_columns:
                        # 사용자가 지정한 컬럼들로 headings 생성
                        headings = []
                        # 무의미한 값 목록 (참조문서 제목으로 부적합)
                        invalid_heading_values = {'-', '--', '없음', 'N/A', 'n/a', 'NA', 'null', 'None', '해당없음', '해당 없음'}
                        for col in mapping.heading_columns:
                            if col in row.data and row.data[col]:
                                val = str(row.data[col]).strip()
                                # 무의미한 값은 제외하고 유효한 값만 추가
                                if val and val not in invalid_heading_values:
                                    headings.append(val)
                        metadata["headings"] = headings if headings else [request.file_name, f"행 {row.row_index + 1}"]
                    else:
                        # 기본값: [파일명, 행 번호]
                        metadata["headings"] = [request.file_name, f"행 {row.row_index + 1}"]

                    metadata_list.append(metadata)

                # Qdrant 업로드
                vector_ids = await qdrant_service.upsert_vectors(
                    collection_name=request.collection_name,
                    vectors=embeddings,
                    texts=texts,
                    metadata_list=metadata_list
                )

                # 결과 기록
                for i, row in enumerate(batch_rows):
                    id_value = None
                    if mapping.id_column and mapping.id_column in row.data:
                        id_value = row.data[mapping.id_column]

                    results.append(DynamicEmbeddingResult(
                        row_index=row.row_index,
                        id_value=id_value,
                        success=True,
                        vector_id=vector_ids[i] if i < len(vector_ids) else None
                    ))
                    success_count += 1

                print(f"[INFO] Embedded batch {batch_start+1}-{batch_end}")

                # SQLite 기록용 데이터 수집
                vector_ids_all.extend(vector_ids)
                texts_all.extend(texts)

            except Exception as e:
                print(f"[ERROR] Failed to embed Excel batch {batch_start+1}-{batch_end}: {e}")
                for row in batch_rows:
                    id_value = None
                    if mapping.id_column and mapping.id_column in row.data:
                        id_value = row.data[mapping.id_column]

                    results.append(DynamicEmbeddingResult(
                        row_index=row.row_index,
                        id_value=id_value,
                        success=False,
                        error=str(e)
                    ))
                    failure_count += 1

        # SQLite에 문서 및 업로드 이력 기록
        if success_count > 0:
            try:
                from backend.services.excel_document_service import (
                    generate_excel_metadata_content,
                    generate_preview,
                    calculate_total_length
                )

                # rows를 dict로 변환 (Pydantic 모델 -> dict)
                rows_dict = [{"row_index": r.row_index, "data": r.data} for r in request.rows]
                mapping_dict = request.mapping.model_dump()

                # Document 생성
                excel_doc = Document(
                    task_id=f"excel-{uuid.uuid4().hex[:12]}",
                    original_filename=request.file_name,
                    file_size=None,
                    file_type="xlsx",
                    status="success",
                    content_length=calculate_total_length(rows_dict, mapping_dict.get("text_columns", [])),
                    content_preview=generate_preview(texts_all[:5]),
                    md_content=generate_excel_metadata_content(
                        request.file_name,
                        rows_dict,
                        mapping_dict,
                        len(request.rows)
                    ),
                    category=request.collection_name,
                    parse_options={
                        "source_type": "excel",
                        "mapping": mapping_dict,
                        "total_rows": len(request.rows)
                    }
                )
                db.add(excel_doc)
                db.flush()  # ID 획득

                # QdrantUploadHistory 생성
                history = QdrantUploadHistory(
                    document_id=excel_doc.id,
                    collection_name=request.collection_name,
                    chunk_count=success_count,
                    vector_ids_json=json.dumps(vector_ids_all),
                    qdrant_url=settings.QDRANT_URL,
                    upload_status="success"
                )
                db.add(history)
                db.commit()

                logger.info(f"Excel document saved to SQLite: id={excel_doc.id}, filename={request.file_name}, collection={request.collection_name}")

            except Exception as db_error:
                logger.error(f"Failed to save Excel document to SQLite: {db_error}")
                db.rollback()
                # Qdrant 업로드는 성공했으므로 에러를 던지지 않고 로그만 기록

        return DynamicEmbeddingResponse(
            total=len(rows),
            success_count=success_count,
            failure_count=failure_count,
            results=results
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Dynamic embedding failed: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Excel 임베딩 실패: {str(e)}"
        )


@router.post("/migrate/excel-collection/{collection_name}")
async def migrate_excel_collection(
    collection_name: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    기존 엑셀 컬렉션을 SQLite에 역으로 기록

    Qdrant에서 메타데이터를 추출하여 Document 및 QdrantUploadHistory 생성
    프롬프트 자동생성 모달에서 엑셀 문서가 표시되도록 함
    """
    try:
        # 1. 컬렉션 존재 확인
        collection_exists = await qdrant_service.collection_exists(collection_name)
        if not collection_exists:
            raise HTTPException(
                status_code=404,
                detail=f"Collection '{collection_name}'이 존재하지 않습니다"
            )

        # 2. 이미 마이그레이션된 문서가 있는지 확인
        existing_history = db.query(QdrantUploadHistory).filter(
            QdrantUploadHistory.collection_name == collection_name
        ).first()
        if existing_history:
            raise HTTPException(
                status_code=400,
                detail=f"Collection '{collection_name}'은 이미 SQLite에 기록되어 있습니다"
            )

        # 3. Qdrant에서 포인트 조회 (scroll)
        all_points = []
        offset = None
        while True:
            points, next_offset = await qdrant_service.client.scroll(
                collection_name=collection_name,
                limit=100,
                offset=offset,
                with_payload=True,
                with_vectors=False
            )
            all_points.extend(points)
            if next_offset is None:
                break
            offset = next_offset

        if not all_points:
            raise HTTPException(
                status_code=400,
                detail=f"Collection '{collection_name}'에 데이터가 없습니다"
            )

        # 4. source_file 기준 그룹핑
        grouped = {}
        for point in all_points:
            payload = point.payload or {}
            source = payload.get("source_file", "unknown.xlsx")
            if source not in grouped:
                grouped[source] = {
                    "texts": [],
                    "vector_ids": [],
                    "row_indices": [],
                    "sample_payload": payload
                }
            grouped[source]["texts"].append(payload.get("text", ""))
            grouped[source]["vector_ids"].append(str(point.id))
            grouped[source]["row_indices"].append(payload.get("row_index", 0))

        # 5. 각 source_file에 대해 Document 생성
        created_docs = []
        for source_file, data in grouped.items():
            from backend.services.excel_document_service import generate_preview

            # 메타정보 생성
            sample = data["sample_payload"]
            text_columns = [k for k in sample.keys() if k not in [
                "source_file", "row_index", "id", "tags", "headings", "text"
            ]]

            md_content = f"""# {source_file}

## 문서 정보
- **유형**: Excel 데이터 (마이그레이션됨)
- **총 행 수**: {len(data["texts"])}
- **컬렉션**: {collection_name}

## 메타데이터 컬럼
- {', '.join(text_columns) if text_columns else '없음'}

## 샘플링 안내
이 문서의 내용은 Qdrant 벡터 DB에 행 단위로 임베딩되어 있습니다.
프롬프트 자동생성 시 청크 기반 샘플링을 통해 의미론적으로 관련된 행들이 자동 추출됩니다.
"""

            doc = Document(
                task_id=f"excel-migrated-{uuid.uuid4().hex[:8]}",
                original_filename=source_file,
                file_type="xlsx",
                status="success",
                content_length=sum(len(t) for t in data["texts"]),
                content_preview=generate_preview(data["texts"][:5]),
                md_content=md_content,
                category=collection_name,
                parse_options={
                    "source_type": "excel",
                    "migrated": True,
                    "original_collection": collection_name
                }
            )
            db.add(doc)
            db.flush()

            history = QdrantUploadHistory(
                document_id=doc.id,
                collection_name=collection_name,
                chunk_count=len(data["texts"]),
                vector_ids_json=json.dumps(data["vector_ids"]),
                qdrant_url=settings.QDRANT_URL,
                upload_status="success"
            )
            db.add(history)
            created_docs.append({
                "document_id": doc.id,
                "filename": source_file,
                "rows": len(data["texts"])
            })

        db.commit()

        logger.info(f"Excel collection migrated: {collection_name}, {len(created_docs)} documents")

        return {
            "success": True,
            "collection_name": collection_name,
            "migrated_documents": len(created_docs),
            "documents": created_docs,
            "total_points": len(all_points)
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to migrate Excel collection: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"마이그레이션 실패: {str(e)}"
        )
