"""
피드백 API 라우터
사용자 피드백 제출 및 조회 엔드포인트
"""

import logging
import io
from datetime import date, datetime
from typing import Optional, List
from urllib.parse import quote

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.cell import WriteOnlyCell

from backend.database import get_db
from backend.dependencies.auth import get_current_active_user
from backend.services.feedback_crud import feedback_crud

logger = logging.getLogger(__name__)


# ============================================================
# Pydantic Schemas
# ============================================================

class FeedbackCreate(BaseModel):
    """피드백 생성 요청"""
    message_id: str = Field(..., description="프론트엔드 메시지 ID")
    session_id: str = Field(..., description="채팅 세션 ID")
    collection_name: str = Field(..., description="컬렉션명")
    rating: str = Field(..., description="평가 (positive | negative)")
    category: Optional[str] = Field(None, description="부정 피드백 카테고리")
    comment: Optional[str] = Field(None, description="추가 의견")
    user_query: str = Field(..., description="사용자 질문")
    assistant_response: Optional[str] = Field(None, description="AI 응답")
    llm_model: Optional[str] = Field(None, description="LLM 모델명")
    reasoning_level: Optional[str] = Field(None, description="추론 레벨")
    retrieved_docs_count: Optional[int] = Field(None, description="참조 문서 수")


class FeedbackResponse(BaseModel):
    """피드백 생성 응답"""
    feedback_id: str
    message_id: str
    rating: str
    created_at: str


class FeedbackSummaryResponse(BaseModel):
    """피드백 요약 응답"""
    total_count: int
    positive_count: int
    negative_count: int
    positive_rate: float
    category_distribution: dict
    daily_trend: List[dict]


class FeedbackListResponse(BaseModel):
    """피드백 목록 응답"""
    feedbacks: List[dict]
    total: int
    skip: int
    limit: int


class FeedbackExistsResponse(BaseModel):
    """피드백 존재 여부 응답"""
    exists: bool
    rating: Optional[str] = None


# ============================================================
# Router
# ============================================================

router = APIRouter(
    prefix="/api/feedback",
    tags=["feedback"],
)


@router.post("/", response_model=FeedbackResponse)
async def create_feedback(
    request: FeedbackCreate,
    db: Session = Depends(get_db),
):
    """
    피드백 제출 (인증 불필요)

    Args:
        request: 피드백 생성 요청
        db: 데이터베이스 세션

    Returns:
        FeedbackResponse: 생성된 피드백 정보
    """
    try:
        # 평가 값 검증
        if request.rating not in ["positive", "negative"]:
            raise HTTPException(
                status_code=400,
                detail="rating must be 'positive' or 'negative'"
            )

        # 부정 피드백 카테고리 검증
        valid_categories = ["inaccurate", "incomplete", "irrelevant", "outdated", "other", None]
        if request.category not in valid_categories:
            raise HTTPException(
                status_code=400,
                detail=f"category must be one of: {valid_categories}"
            )

        feedback = await feedback_crud.create_feedback(
            db=db,
            message_id=request.message_id,
            session_id=request.session_id,
            collection_name=request.collection_name,
            rating=request.rating,
            category=request.category,
            comment=request.comment,
            user_query=request.user_query,
            assistant_response=request.assistant_response,
            llm_model=request.llm_model,
            reasoning_level=request.reasoning_level,
            retrieved_docs_count=request.retrieved_docs_count,
        )

        return FeedbackResponse(
            feedback_id=feedback.feedback_id,
            message_id=feedback.message_id,
            rating=feedback.rating,
            created_at=feedback.created_at.isoformat() if feedback.created_at else "",
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"피드백 생성 실패: {e}")
        raise HTTPException(status_code=500, detail=f"피드백 생성 실패: {str(e)}")


@router.get("/check/{message_id}", response_model=FeedbackExistsResponse)
async def check_feedback_exists(
    message_id: str,
    db: Session = Depends(get_db),
):
    """
    메시지에 대한 피드백 존재 여부 확인 (인증 불필요)

    Args:
        message_id: 메시지 ID
        db: 데이터베이스 세션

    Returns:
        FeedbackExistsResponse: 피드백 존재 여부
    """
    try:
        rating = await feedback_crud.check_feedback_exists(db, message_id)
        return FeedbackExistsResponse(
            exists=rating is not None,
            rating=rating,
        )
    except Exception as e:
        logger.error(f"피드백 확인 실패: {e}")
        raise HTTPException(status_code=500, detail=f"피드백 확인 실패: {str(e)}")


@router.get("/summary", response_model=FeedbackSummaryResponse, dependencies=[Depends(get_current_active_user)])
async def get_feedback_summary(
    collection_name: Optional[str] = Query(None, description="컬렉션 필터"),
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    db: Session = Depends(get_db),
):
    """
    피드백 요약 통계 조회 (관리자 전용)

    Args:
        collection_name: 컬렉션 필터 (선택)
        date_from: 시작 날짜 (선택)
        date_to: 종료 날짜 (선택)
        db: 데이터베이스 세션

    Returns:
        FeedbackSummaryResponse: 피드백 요약 통계
    """
    try:
        # "ALL" 처리
        effective_collection = None if collection_name in (None, "ALL") else collection_name

        summary = await feedback_crud.get_feedback_summary(
            db=db,
            collection_name=effective_collection,
            date_from=date_from,
            date_to=date_to,
        )
        return FeedbackSummaryResponse(**summary)

    except Exception as e:
        logger.error(f"피드백 요약 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=f"피드백 요약 조회 실패: {str(e)}")


@router.get("/list", response_model=FeedbackListResponse, dependencies=[Depends(get_current_active_user)])
async def get_feedback_list(
    collection_name: Optional[str] = Query(None, description="컬렉션 필터"),
    rating: Optional[str] = Query(None, description="평가 필터 (positive/negative)"),
    category: Optional[str] = Query(None, description="카테고리 필터"),
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    skip: int = Query(0, ge=0, description="건너뛸 항목 수"),
    limit: int = Query(20, ge=1, le=100, description="조회할 항목 수"),
    db: Session = Depends(get_db),
):
    """
    피드백 목록 조회 (관리자 전용)

    Args:
        collection_name: 컬렉션 필터 (선택)
        rating: 평가 필터 (선택)
        category: 카테고리 필터 (선택)
        date_from: 시작 날짜 (선택)
        date_to: 종료 날짜 (선택)
        skip: 건너뛸 항목 수
        limit: 조회할 항목 수
        db: 데이터베이스 세션

    Returns:
        FeedbackListResponse: 피드백 목록
    """
    try:
        # "ALL" 처리
        effective_collection = None if collection_name in (None, "ALL") else collection_name

        result = await feedback_crud.get_feedback_list(
            db=db,
            collection_name=effective_collection,
            rating=rating,
            category=category,
            date_from=date_from,
            date_to=date_to,
            skip=skip,
            limit=limit,
        )
        return FeedbackListResponse(**result)

    except Exception as e:
        logger.error(f"피드백 목록 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=f"피드백 목록 조회 실패: {str(e)}")


@router.get("/recent-negative", dependencies=[Depends(get_current_active_user)])
async def get_recent_negative_feedbacks(
    collection_name: Optional[str] = Query(None, description="컬렉션 필터"),
    limit: int = Query(10, ge=1, le=50, description="조회할 항목 수"),
    db: Session = Depends(get_db),
):
    """
    최근 부정 피드백 조회 (관리자 전용)

    Args:
        collection_name: 컬렉션 필터 (선택)
        limit: 조회할 항목 수
        db: 데이터베이스 세션

    Returns:
        List[dict]: 최근 부정 피드백 목록
    """
    try:
        # "ALL" 처리
        effective_collection = None if collection_name in (None, "ALL") else collection_name

        feedbacks = await feedback_crud.get_recent_negative_feedbacks(
            db=db,
            collection_name=effective_collection,
            limit=limit,
        )
        return {"feedbacks": feedbacks, "total": len(feedbacks)}

    except Exception as e:
        logger.error(f"최근 부정 피드백 조회 실패: {e}")
        raise HTTPException(status_code=500, detail=f"최근 부정 피드백 조회 실패: {str(e)}")


@router.get("/export/excel", dependencies=[Depends(get_current_active_user)])
async def export_feedback_to_excel(
    collection_name: Optional[str] = Query(None, description="컬렉션 필터"),
    date_from: Optional[date] = Query(None, description="시작 날짜"),
    date_to: Optional[date] = Query(None, description="종료 날짜"),
    db: Session = Depends(get_db),
):
    """
    피드백을 Excel 파일로 내보내기 (관리자 전용)

    Args:
        collection_name: 컬렉션 필터 (선택)
        date_from: 시작 날짜 (선택)
        date_to: 종료 날짜 (선택)
        db: 데이터베이스 세션

    Returns:
        StreamingResponse: Excel 파일
    """
    try:
        # 날짜 기본값 설정
        if not date_from:
            date_from = date.today()
        if not date_to:
            date_to = date.today()

        # "ALL" 처리
        effective_collection = None if collection_name in (None, "ALL") else collection_name

        # 피드백 조회 (전체)
        result = await feedback_crud.get_feedback_list(
            db=db,
            collection_name=effective_collection,
            date_from=date_from,
            date_to=date_to,
            skip=0,
            limit=10000,
        )

        feedbacks = result["feedbacks"]

        # Excel 워크북 생성
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="피드백")

        # 헤더 정의
        headers = [
            "날짜/시간", "컬렉션", "평가", "카테고리", "사용자 의견",
            "사용자 질문", "AI 응답", "LLM 모델", "추론 레벨", "참조문서 수"
        ]

        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)

        # 헤더 행 작성
        header_row = []
        for header in headers:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            header_row.append(cell)
        ws.append(header_row)

        # 카테고리 한글 매핑
        category_map = {
            "inaccurate": "부정확함",
            "incomplete": "불완전함",
            "irrelevant": "관련없음",
            "outdated": "구버전",
            "other": "기타",
            None: "",
        }

        # 데이터 작성
        for fb in feedbacks:
            data_row = []
            row_data = {
                "날짜/시간": fb.get("created_at", ""),
                "컬렉션": fb.get("collection_name", ""),
                "평가": "긍정" if fb.get("rating") == "positive" else "부정",
                "카테고리": category_map.get(fb.get("category"), fb.get("category", "")),
                "사용자 의견": fb.get("comment", ""),
                "사용자 질문": fb.get("user_query", ""),
                "AI 응답": fb.get("assistant_response", ""),
                "LLM 모델": fb.get("llm_model", ""),
                "추론 레벨": fb.get("reasoning_level", ""),
                "참조문서 수": fb.get("retrieved_docs_count", ""),
            }
            for header in headers:
                value = row_data.get(header, "")
                cell = WriteOnlyCell(ws, value=value)
                cell.alignment = cell_alignment
                data_row.append(cell)
            ws.append(data_row)

        # 메모리 스트림에 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 파일명 생성
        filename = f"feedback_{collection_name or 'ALL'}_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
        encoded_filename = quote(filename, safe='')

        logger.info(f"피드백 Excel 다운로드: {len(feedbacks)}건, 기간: {date_from} ~ {date_to}")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except Exception as e:
        logger.error(f"피드백 Excel 내보내기 실패: {e}")
        raise HTTPException(status_code=500, detail=f"피드백 Excel 내보내기 실패: {str(e)}")
